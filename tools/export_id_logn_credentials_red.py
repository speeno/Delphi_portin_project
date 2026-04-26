#!/usr/bin/env python3
"""RED — 운영 Id_Logn 로그인 ID/PW 추출 도구.

본 스크립트는 운영자 본인의 자격증명 운반/재생성 작업을 위한 RED 등급 도구입니다.
정책 정합:
  * 출력 파일은 `WeLove_FTP/` 하위로만 작성합니다 — 해당 폴더는 `.gitignore` 로
    전체 제외되어 있어 git 추적·원격 푸시 위험이 없습니다.
  * 스크립트 자체에는 자격증명을 한 줄도 포함하지 않습니다 (소스에는 0건).
  * `docs/secrets-policy.md` SEC-POL-STORAGE-01 의 (1) 작업자 PC `WeLove_FTP/...`
    (gitignored) 허용 위치를 사용합니다.

사용:
    python3 tools/export_id_logn_credentials_red.py --dry-run
    python3 tools/export_id_logn_credentials_red.py --server remote_153 --output WeLove_FTP/_oneoff_distributor_credentials_DELETE_AFTER_USE.xlsx

사용 후:
    1. 운영자가 vault / 1Password 로 옮긴 즉시
       `WeLove_FTP/_oneoff_distributor_credentials_DELETE_AFTER_USE.xlsx`
       를 영구 삭제하십시오.
    2. 영구 삭제는 finder 휴지통이 아니라 `rm -P` 또는 디스크 zeroize 권장.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as _dt
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent

DB_INFO_XLSX = (
    ROOT
    / "WeLove_FTP/Welove_인수인계/셋팅방법/DB정보, DB_FTP 엑셀/DB정보 엑셀정리.xlsx"
)
DB_FTP_XLSX = (
    ROOT
    / "WeLove_FTP/Welove_인수인계/셋팅방법/DB정보, DB_FTP 엑셀/DB_FTP 엑셀정리.xlsx"
)

OUT_PATH = ROOT / "WeLove_FTP" / "_oneoff_distributor_credentials_DELETE_AFTER_USE.xlsx"
BACKEND = ROOT / "도서물류관리프로그램" / "backend"
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

from app.core.config import get_server_profiles  # noqa: E402
from app.core.db import close_all_pools, execute_query  # noqa: E402
from app.services import auth_service  # noqa: E402

# tools/seed_tenants_directory.py 의 T2_DIST 정의와 일치
_PRIMARY_DIST_FAMILIES = {f"chul_{i:02d}" for i in range(1, 9)} | {"book_kb", "kb_book"}

# 운영상 별도 분류이지만 "총판/물류" 라벨이 붙어있는 보조 후보
# (s1_chul_8x 시리즈, sky_01 등 — 운영자가 사후 판단)
_SECONDARY_DIST_FAMILIES = {
    "chul_83",
    "chul_84",
    "chul_85",
    "chul_87",
    "chul_88",
    "chul_42",
}

_DIST_KEYWORDS = ("물류", "도서유통", "북앤더", "북스로드", "한강북", "한강도서라인")
_DB_NAME_RE = re.compile(r"^[A-Za-z0-9_\-]{1,64}$")
_SERVER_LABEL_TO_REMOTE = {
    "서버1": "remote_154",
    "서버2": "remote_155",
    "서버3": "remote_153",
    "서버4": "remote_138",
}
_LIVE_SHEET_HEADERS = [
    "서버",
    "DB 명",
    "account_family",
    "tenant/DB 표시명",
    "계정 유형",
    "build_role",
    "warehouse_menu_tier",
    "dist_hcode",
    "Hcode",
    "Hname",
    "로그인 ID (=Id_Logn.Gcode)",
    "로그인 PW (=Id_Logn.Gpass)",
    "Gname",
    "source",
]
DEFAULT_EXCLUDED_HCODES = {"0000"}


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_family_from_folder(folder: str) -> str:
    """`s2_chul_04` → `chul_04`, `s4_book_kb` → `book_kb`, `book_07` → `book_07`."""
    folder = folder.strip()
    if not folder:
        return ""
    if folder.startswith(("s1_", "s2_", "s3_", "s4_", "sx_", "sy_")):
        return folder.split("_", 1)[1]
    return folder


def _remote_id(server_label: str) -> str:
    raw = (server_label or "").strip()
    if raw in _SERVER_LABEL_TO_REMOTE:
        return _SERVER_LABEL_TO_REMOTE[raw]
    if raw.startswith("remote_"):
        return raw
    if "(" in raw and ")" in raw:
        inner = raw.split("(", 1)[1].split(")", 1)[0].strip()
        return _SERVER_LABEL_TO_REMOTE.get(inner, inner)
    return raw


def _safe_db_name(db_name: str) -> str:
    dbn = _norm(db_name).rstrip()
    return dbn if _DB_NAME_RE.match(dbn) else ""


def _infer_family_from_db_name(db_name: str) -> str:
    dbn = _norm(db_name)
    if dbn.endswith("_db"):
        return dbn[:-3]
    return dbn


def _is_distributor(family: str, tenant_name: str) -> tuple[bool, str]:
    """Return (is_dist, classification)."""
    if family in _PRIMARY_DIST_FAMILIES:
        return True, "T2_DIST (정본)"
    if family in _SECONDARY_DIST_FAMILIES:
        return True, "T2_DIST (보조/legacy)"
    if any(kw in tenant_name for kw in _DIST_KEYWORDS):
        return True, "T2_DIST (이름 키워드 일치)"
    return False, ""


def _load_db_info() -> dict[tuple[str, str], dict]:
    """Load DB정보 엑셀정리.xlsx → {(server_id, account_family): {db_user, db_pass, db_name, server_label}}"""
    wb = openpyxl.load_workbook(str(DB_INFO_XLSX), data_only=True)
    sh = wb["Sheet1"]
    out: dict[tuple[str, str], dict] = {}
    for row in sh.iter_rows(values_only=True):
        cells = [_norm(v) for v in row]
        if len(cells) < 7:
            continue
        # 헤더/빈 행 skip
        if "서버" in cells and "물류사명(업체명)" in cells:
            continue
        server_label, tenant, db_user, db_pass, db_name = (
            cells[2], cells[3], cells[4], cells[5], cells[6],
        )
        if not server_label or not tenant or not db_user:
            continue
        sid = server_label.split("(", 1)[1].rstrip(")") if "(" in server_label else server_label
        family = db_user[:-5] if db_user.endswith("_user") else db_user.split("_user", 1)[0]
        out[(sid, family)] = {
            "server_label": server_label,
            "server_id": sid,
            "tenant_name_kor": tenant,
            "account_family": family,
            "db_user": db_user,
            "db_password": db_pass,
            "db_name": db_name,
        }
    return out


def _load_db_ftp_rows() -> list[dict]:
    """Load DB_FTP 엑셀정리.xlsx → list of program login rows."""
    wb = openpyxl.load_workbook(str(DB_FTP_XLSX), data_only=True)
    sh = wb["Sheet1"]
    rows_out: list[dict] = []
    for row in sh.iter_rows(values_only=True):
        cells = [_norm(v) for v in row]
        if len(cells) < 11:
            continue
        if "서버" in cells and "물류사명(업체명)" in cells:
            continue
        server_label, folder, tenant_name = cells[2], cells[3], cells[4]
        ftp_user_a, ftp_user_b = cells[5], cells[6]
        mulryu_code = cells[7]
        program_user, login_id, login_pw = cells[8], cells[9], cells[10]
        if not folder or not server_label:
            continue
        family = _parse_family_from_folder(folder)
        if not family:
            continue
        if not (tenant_name or login_id):
            continue
        sid = server_label.split("(", 1)[1].rstrip(")") if "(" in server_label else server_label
        rows_out.append(
            {
                "server_label": server_label,
                "server_id": sid,
                "folder": folder,
                "account_family": family,
                "tenant_name_kor": tenant_name,
                "ftp_user_main": ftp_user_a,
                "ftp_user_sub": ftp_user_b,
                "mulryu_code": mulryu_code,
                "program_user": program_user,
                "login_id": login_id,
                "login_pw": login_pw,
            }
        )
    return rows_out


def _build_db_candidate_index(
    db_info_idx: dict[tuple[str, str], dict],
    ftp_rows: list[dict],
) -> dict[tuple[str, str], dict[str, Any]]:
    out: dict[tuple[str, str], dict[str, Any]] = {}
    for db in db_info_idx.values():
        remote_id = _remote_id(db.get("server_id", ""))
        db_name = _safe_db_name(db.get("db_name", ""))
        family = _norm(db.get("account_family"))
        if not remote_id or not db_name:
            continue
        out[(remote_id, db_name)] = {
            "server_id": remote_id,
            "server_label": db.get("server_id", ""),
            "db_name": db_name,
            "account_family": family,
            "tenant_name_kor": db.get("tenant_name_kor", ""),
            "db_user": db.get("db_user", ""),
            "db_password": db.get("db_password", ""),
        }
    for fr in ftp_rows:
        db = db_info_idx.get((fr["server_id"], fr["account_family"]), {})
        remote_id = _remote_id(fr.get("server_id", ""))
        db_name = _safe_db_name(db.get("db_name", ""))
        if not remote_id or not db_name:
            continue
        out.setdefault(
            (remote_id, db_name),
            {
                "server_id": remote_id,
                "server_label": fr.get("server_id", ""),
                "db_name": db_name,
                "account_family": fr.get("account_family", ""),
                "tenant_name_kor": fr.get("tenant_name_kor", ""),
                "db_user": db.get("db_user", ""),
                "db_password": db.get("db_password", ""),
            },
        )
    return out


def _db_info_by_remote_db(db_info_idx: dict[tuple[str, str], dict]) -> dict[tuple[str, str], dict[str, Any]]:
    out: dict[tuple[str, str], dict[str, Any]] = {}
    for db in db_info_idx.values():
        remote_id = _remote_id(db.get("server_id", ""))
        db_name = _safe_db_name(db.get("db_name", ""))
        if not remote_id or not db_name:
            continue
        out[(remote_id, db_name)] = {
            "server_id": remote_id,
            "db_name": db_name,
            "account_family": _norm(db.get("account_family")) or _infer_family_from_db_name(db_name),
            "tenant_name_kor": db.get("tenant_name_kor", ""),
            "db_user": db.get("db_user", ""),
            "db_password": db.get("db_password", ""),
        }
    return out


async def _list_server_databases(server_id: str) -> list[str]:
    try:
        rows = await execute_query(server_id, "SHOW DATABASES")
    except Exception:
        return []
    out: list[str] = []
    for r in rows or []:
        if not isinstance(r, dict) or not r:
            continue
        dbn = _safe_db_name(str(next(iter(r.values()))))
        if dbn:
            out.append(dbn)
    return sorted(set(out))


async def _build_live_db_candidates(
    db_info_idx: dict[tuple[str, str], dict],
    *,
    server_filter: set[str] | None = None,
) -> dict[tuple[str, str], dict[str, Any]]:
    info_by_db = _db_info_by_remote_db(db_info_idx)
    out: dict[tuple[str, str], dict[str, Any]] = {}
    for profile in get_server_profiles():
        server_id = _norm(profile.get("id"))
        if not server_id:
            continue
        if server_filter and server_id not in server_filter:
            continue
        for db_name in await _list_server_databases(server_id):
            key = (server_id, db_name)
            meta = info_by_db.get(key, {})
            out[key] = {
                "server_id": server_id,
                "server_label": profile.get("label", ""),
                "db_name": db_name,
                "account_family": meta.get("account_family") or _infer_family_from_db_name(db_name),
                "tenant_name_kor": meta.get("tenant_name_kor", ""),
                "db_user": meta.get("db_user", ""),
                "db_password": meta.get("db_password", ""),
            }
    return out


def _row_has_id_logn_table(row: dict[str, Any]) -> bool:
    return any(str(v).strip().lower() == "id_logn" for v in row.values())


async def _db_has_id_logn(server_id: str, db_name: str) -> bool:
    try:
        rows = await execute_query(server_id, f"SHOW TABLES FROM `{db_name}`")
    except Exception:
        return False
    return any(_row_has_id_logn_table(r) for r in rows or [] if isinstance(r, dict))


async def _load_live_id_logn_rows(
    candidates: dict[tuple[str, str], dict[str, Any]],
    *,
    excluded_hcodes: set[str],
) -> list[dict[str, Any]]:
    live_rows: list[dict[str, Any]] = []
    for _, cand in sorted(candidates.items()):
        server_id = cand["server_id"]
        db_name = cand["db_name"]
        if not await _db_has_id_logn(server_id, db_name):
            continue
        sql = (
            "SELECT gcode AS gcode, hcode AS hcode, hname AS hname, "
            "gname AS gname, gpass AS gpass "
            f"FROM `{db_name}`.Id_Logn "
            "ORDER BY hcode, gcode"
        )
        try:
            rows = await execute_query(server_id, sql)
        except Exception:
            continue
        for r in rows or []:
            if not isinstance(r, dict):
                continue
            gcode = _norm(r.get("gcode"))
            hcode = _norm(r.get("hcode"))
            if not gcode and not hcode:
                continue
            if hcode in excluded_hcodes:
                continue
            acct_meta = auth_service._resolve_account_type(
                gcode,
                hcode,
                server_id,
                account_family=cand.get("account_family") or None,
                resolved_db=db_name,
            )
            live_rows.append(
                {
                    "server_id": server_id,
                    "db_name": db_name,
                    "account_family": cand.get("account_family", ""),
                    "tenant_name_kor": cand.get("tenant_name_kor", ""),
                    "account_type": acct_meta.get("account_type") or "",
                    "build_role": acct_meta.get("build_role") or "",
                    "warehouse_menu_tier": acct_meta.get("warehouse_menu_tier") or "",
                    "dist_hcode": acct_meta.get("dist_hcode") or "",
                    "hcode": hcode,
                    "hname": _norm(r.get("hname")),
                    "gcode": gcode,
                    "gpass": _norm(r.get("gpass")),
                    "gname": _norm(r.get("gname")),
                    "source": "live Id_Logn",
                }
            )
    return live_rows


def _safe_sheet_title(raw: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]\:\*\?\/\\]", "_", raw).strip() or "HCODE"
    title = title[:31]
    base = title
    n = 2
    while title in used:
        suffix = f"_{n}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        n += 1
    used.add(title)
    return title


def _write_warning_sheet(wb: openpyxl.Workbook) -> None:
    sh = wb.create_sheet("⚠ 사용 전 필독", 0)
    today = _dt.date.today().isoformat()
    delete_by = (_dt.date.today() + _dt.timedelta(days=7)).isoformat()
    msgs = [
        ("⚠️ 본 파일은 1회성 운반용입니다 — 사용 즉시 영구 삭제 ⚠️", True, "FFC7CE"),
        ("", False, None),
        (f"생성일: {today}", False, None),
        (f"삭제 권장일: {delete_by} (생성일 +7일 이내)", True, None),
        ("", False, None),
        ("정책 (docs/secrets-policy.md)", True, "BDD7EE"),
        ("- SEC-POL-CLASS-RED — 본 파일은 RED 등급(자격증명 평문 포함)입니다.", False, None),
        ("- SEC-POL-STORAGE-01 — 허용 저장 위치만 사용 (WeLove_FTP/ 는 gitignored).", False, None),
        ("- SEC-POL-STORAGE-04 — 외부 클라우드(Drive·Dropbox 등) 업로드 금지.", False, None),
        ("- 회의·PR·이슈·메신저 본문 평문 인용 금지 (마스킹 메타만 허용).", False, None),
        ("", False, None),
        ("사용 절차", True, "BDD7EE"),
        ("1. vault / 1Password / 사내 secret store 로 즉시 이관.", False, None),
        ("2. 운영 환경의 .env / CI secret 으로 주입.", False, None),
        ("3. 본 xlsx 를 `rm -P` (또는 디스크 zeroize) 로 영구 삭제.", False, None),
        ("4. tools/export_id_logn_credentials_red.py 는 RED 운영 도구로 유지하되, xlsx 는 남기지 않습니다.", False, None),
        ("", False, None),
        ("출처", True, "BDD7EE"),
        ("- 원본 1: WeLove_FTP/Welove_인수인계/셋팅방법/DB정보, DB_FTP 엑셀/DB정보 엑셀정리.xlsx (MAN-017)", False, None),
        ("- 원본 2: WeLove_FTP/Welove_인수인계/셋팅방법/DB정보, DB_FTP 엑셀/DB_FTP 엑셀정리.xlsx (MAN-018)", False, None),
        ("- 원본 3: 운영 DB 각 스키마의 Id_Logn 라이브 조회 (서버별 시트, Hcode=0000 제외)", False, None),
        ("- 분류 기준: tools/seed_tenants_directory.py _PRIMARY_DIST_FAMILIES (T2_DIST 정본)", False, None),
        ("", False, None),
        ("통합 로그인(웹 API) vs DB 접속 계정", True, "BDD7EE"),
        (
            "- 시트「총판(T2_DIST) 로그인 ID·PW」에서: 「로그인 ID」= Id_Logn.Gcode, 「로그인 PW」= Id_Logn.Gpass "
            "(POST /api/v1/auth/login 의 userId / password).",
            False,
            None,
        ),
        (
            "- 같은 시트의 「DB 사용자」「DB 비번」은 MySQL 서버 접속용 계정으로, Id_Logn 과 별개입니다.",
            False,
            None,
        ),
        ("- 상세 매핑 표는 시트「필드_매핑_Id_Logn」을 참고하세요.", False, None),
        ("- 서버별 시트는 운영 DB Id_Logn.Gcode/Gpass를 그대로 담되, Hcode=0000 계정은 제외합니다.", False, None),
        ("", False, None),
        ("위반 시", True, "FFC7CE"),
        ("- git 추적 발견 → 즉시 history rewrite + 모든 자격증명 회전 (정책 §6).", False, None),
    ]
    for ri, (text, bold, fill_hex) in enumerate(msgs, start=1):
        cell = sh.cell(row=ri, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            cell.font = Font(bold=True, size=12)
        if fill_hex:
            cell.fill = PatternFill("solid", fgColor=fill_hex)
    sh.column_dimensions["A"].width = 110


def _write_field_mapping_sheet(wb: openpyxl.Workbook) -> None:
    """통합 로그인 필드(Id_Logn) vs DB 접속 계정 구분 — 검증용 문서."""
    sh = wb.create_sheet("필드_매핑_Id_Logn", 1)
    rows = [
        ("구분", "본 엑셀 컬럼명", "레거시 DB / API", "비고"),
        (
            "웹 로그인 ID",
            "로그인 ID (=Id_Logn.Gcode)",
            "Id_Logn.Gcode",
            "API: LoginRequest.userId (gcode AS user_id)",
        ),
        (
            "웹 로그인 암호",
            "로그인 PW (=Id_Logn.Gpass)",
            "Id_Logn.Gpass",
            "API: LoginRequest.password (gpass AS password)",
        ),
        ("DB 접속 ID", "DB 사용자", "MySQL 사용자명", "프로그램·터널 DB 연결 — Id_Logn 아님"),
        ("DB 접속 암호", "DB 비번", "MySQL 비밀번호", "위와 쌍"),
        ("표시명(참고)", "프로그램 사용자", "Id_Logn.Gname 근처", "로그인 ID와 혼동 금지 — 표시/작업자명"),
        ("회사코드 힌트", "물류사코드", "Id_Logn.Hcode 등", "통합 로그인 옵션 hcode — 동명 ID 분기용"),
    ]
    for ri, tup in enumerate(rows, start=1):
        for ci, val in enumerate(tup, start=1):
            c = sh.cell(row=ri, column=ci, value=val)
            if ri == 1:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="E2EFDA")
    for ci in range(1, 5):
        sh.column_dimensions[get_column_letter(ci)].width = 36


def _write_live_summary_sheet(wb: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    sh = wb.create_sheet("서버별_요약")
    headers = ["서버", "계정 행 수", "DB 수", "Hcode 수", "계정 유형", "build_role", "시트명"]
    for ci, h in enumerate(headers, start=1):
        c = sh.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")

    by_server: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_server[_norm(r.get("server_id")) or "(blank)"].append(r)
    for ri, (server_id, bucket) in enumerate(sorted(by_server.items()), start=2):
        dbs = {_norm(r.get("db_name")) for r in bucket if _norm(r.get("db_name"))}
        hcodes = {_norm(r.get("hcode")) for r in bucket if _norm(r.get("hcode"))}
        account_types = sorted({_norm(r.get("account_type")) for r in bucket if _norm(r.get("account_type"))})
        build_roles = sorted({_norm(r.get("build_role")) for r in bucket if _norm(r.get("build_role"))})
        sheet_name = bucket[0].get("_sheet_name", "")
        vals = [
            server_id,
            len(bucket),
            len(dbs),
            len(hcodes),
            ", ".join(account_types),
            ", ".join(build_roles),
            sheet_name,
        ]
        for ci, val in enumerate(vals, start=1):
            sh.cell(row=ri, column=ci, value=val)
    for ci in range(1, len(headers) + 1):
        sh.column_dimensions[get_column_letter(ci)].width = 22


def _write_server_live_sheets(wb: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    by_server: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_server[_norm(r.get("server_id")) or "(blank)"].append(r)

    used = set(wb.sheetnames)
    for server_id, bucket in sorted(by_server.items()):
        title = _safe_sheet_title(f"{server_id}_Id_Logn", used)
        for r in bucket:
            r["_sheet_name"] = title
        sh = wb.create_sheet(title)
        for ci, h in enumerate(_LIVE_SHEET_HEADERS, start=1):
            c = sh.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E2EFDA")
        keys = [
            "server_id",
            "db_name",
            "account_family",
            "tenant_name_kor",
            "account_type",
            "build_role",
            "warehouse_menu_tier",
            "dist_hcode",
            "hcode",
            "hname",
            "gcode",
            "gpass",
            "gname",
            "source",
        ]
        for ri, r in enumerate(bucket, start=2):
            for ci, key in enumerate(keys, start=1):
                sh.cell(row=ri, column=ci, value=r.get(key, ""))
        for ci in range(1, len(_LIVE_SHEET_HEADERS) + 1):
            sh.column_dimensions[get_column_letter(ci)].width = 22


def _write_dist_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    sh = wb.create_sheet("총판(T2_DIST) 로그인 ID·PW")
    headers = [
        "분류",
        "tenant_id 후보",
        "물류사명(한글)",
        "account_family",
        "서버",
        "폴더",
        "물류사코드",
        "프로그램 사용자",
        "로그인 ID (=Id_Logn.Gcode)",
        "로그인 PW (=Id_Logn.Gpass)",
        "DB 사용자",
        "DB 비번",
        "DB 명",
        "FTP 사용자(메인)",
        "FTP 사용자(보조)",
        "원본행",
    ]
    for ci, h in enumerate(headers, start=1):
        c = sh.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for ri, r in enumerate(rows, start=2):
        for ci, key in enumerate(
            [
                "_classification",
                "_tenant_id_hint",
                "tenant_name_kor",
                "account_family",
                "server_id",
                "folder",
                "mulryu_code",
                "program_user",
                "login_id",
                "login_pw",
                "db_user",
                "db_password",
                "db_name",
                "ftp_user_main",
                "ftp_user_sub",
                "_source_row",
            ],
            start=1,
        ):
            sh.cell(row=ri, column=ci, value=r.get(key, ""))
    for ci in range(1, len(headers) + 1):
        sh.column_dimensions[get_column_letter(ci)].width = 22


async def _build_workbook(
    *,
    dry_run: bool = False,
    server_filter: set[str] | None = None,
    excluded_hcodes: set[str] | None = None,
    output_path: Path = OUT_PATH,
) -> int:
    if not DB_INFO_XLSX.exists() or not DB_FTP_XLSX.exists():
        print("[ERR] 원본 엑셀 누락:", DB_INFO_XLSX, DB_FTP_XLSX)
        return 2

    db_info_idx = _load_db_info()  # (sid, family) → {db_user, db_password, db_name, ...}
    ftp_rows = _load_db_ftp_rows()
    excluded_hcodes = excluded_hcodes if excluded_hcodes is not None else set(DEFAULT_EXCLUDED_HCODES)
    db_candidates = await _build_live_db_candidates(db_info_idx, server_filter=server_filter)
    live_rows = await _load_live_id_logn_rows(db_candidates, excluded_hcodes=excluded_hcodes)

    merged: list[dict] = []
    for ri, fr in enumerate(ftp_rows, start=1):
        family = fr["account_family"]
        is_dist, classification = _is_distributor(family, fr["tenant_name_kor"])
        if not is_dist:
            continue

        db = db_info_idx.get((fr["server_id"], family), {})
        merged.append(
            {
                **fr,
                "db_user": db.get("db_user", ""),
                "db_password": db.get("db_password", ""),
                "db_name": db.get("db_name", ""),
                "_classification": classification,
                "_tenant_id_hint": f"tendir-{family}-{fr['server_id'].lower()}",
                "_source_row": ri,
            }
        )

    if dry_run:
        by_server: dict[str, int] = {}
        by_db = {(r["server_id"], r["db_name"]) for r in live_rows}
        for r in live_rows:
            key = _norm(r.get("server_id")) or "(blank)"
            by_server[key] = by_server.get(key, 0) + 1
        print("[DRY-RUN] xlsx는 쓰지 않았습니다.")
        print(f"[DRY-RUN] Id_Logn DB {len(by_db)}개, 서버 {len(by_server)}개, 계정 행 {len(live_rows)}건")
        for sid, cnt in sorted(by_server.items()):
            print(f"[DRY-RUN]   - {sid}: {cnt}건")
        print(f"[DRY-RUN] 총판 후보 시트 행 {len(merged)}건")
        return 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    # 기본 빈 시트 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_warning_sheet(wb)
    _write_field_mapping_sheet(wb)
    _write_dist_sheet(wb, merged)
    _write_server_live_sheets(wb, live_rows)
    _write_live_summary_sheet(wb, live_rows)
    wb.save(str(output_path))

    # === 자격증명을 절대 콘솔에 출력하지 않습니다 (G3) ===
    by_class: dict[str, int] = {}
    for r in merged:
        by_class[r["_classification"]] = by_class.get(r["_classification"], 0) + 1

    print(f"[OK] 출력: {output_path.relative_to(ROOT)}")
    print(f"     (gitignored 위치 — git 추적 0건)")
    print(f"[총계] 총판 후보 {len(merged)} 건")
    print(f"[총계] 라이브 Id_Logn 계정 {len(live_rows)} 건")
    for cls, cnt in sorted(by_class.items()):
        print(f"        - {cls}: {cnt}")
    print("[중요] vault / 1Password 이관 후 xlsx를 영구 삭제하십시오.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="RED Id_Logn credential workbook exporter")
    parser.add_argument("--dry-run", action="store_true", help="xlsx를 쓰지 않고 행 수만 확인")
    parser.add_argument(
        "--server",
        action="append",
        default=[],
        help="스캔할 server_id. 여러 번 지정 가능. 미지정 시 전체 서버",
    )
    parser.add_argument(
        "--exclude-hcode",
        action="append",
        default=["0000"],
        help="제외할 Hcode. 여러 번 지정 가능. 기본값: 0000",
    )
    parser.add_argument(
        "--output",
        default=str(OUT_PATH.relative_to(ROOT)),
        help="출력 xlsx 경로. 기본값: WeLove_FTP/_oneoff_distributor_credentials_DELETE_AFTER_USE.xlsx",
    )
    args = parser.parse_args()

    async def _run() -> int:
        try:
            output_path = Path(args.output)
            if not output_path.is_absolute():
                output_path = ROOT / output_path
            return await _build_workbook(
                dry_run=args.dry_run,
                server_filter={_norm(v) for v in args.server if _norm(v)} or None,
                excluded_hcodes={_norm(v) for v in args.exclude_hcode if _norm(v)},
                output_path=output_path,
            )
        finally:
            await close_all_pools()

    return asyncio.run(_run())


if __name__ == "__main__":
    raise SystemExit(main())
