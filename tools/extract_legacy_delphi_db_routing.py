#!/usr/bin/env python3
"""레거시 Delphi(도서유통) Chul.pas 내 MySQL 연결 지정 방식 스캔 → 엑셀.

조사 범위
---------
- ``WeLove_FTP/도서유통-New/**/Chul.pas``
- ``WeLove_FTP/도서유통-출판/**/Chul.pas``

추출
----
- ``Base10.Database.(Host|Login|Password|Database)`` 활성 대입(주석 블록 ``{`` ``}`` 밖)
- ``MimeDecodeString('...')`` → base64 디코드 (호스트·DB명·*_user / *_pw 패턴만 — 실 운영 root 비번은 소스에 없음)
- ``Config.Ini`` 의 ``[Client]`` 읽기 여부 (Base/Pcip/Port 등 — DB 4종과 별도)

출력
----
- ``WeLove_FTP/legacy_delphi_db_routing_from_chul_pas.xlsx`` (gitignored 루트)

비밀 정책
---------
- 엑셀에 **운영 root 비밀번호·MAN-017 실비번** 은 넣지 않는다.
- ``*_user`` / ``*_pw`` / ``*_db`` 는 명명 규칙(레거시 빌드 SKU)으로서 기록.
"""

from __future__ import annotations

import base64
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
WELOVE = ROOT / "WeLove_FTP"
OUT_XLSX = WELOVE / "legacy_delphi_db_routing_from_chul_pas.xlsx"

GLOBS = (
    "도서유통-New/**/Chul.pas",
    "도서유통-출판/**/Chul.pas",
)

ASSIGN_RE = re.compile(
    r"^\s*Base10\.Database\.(Host|Login|Password|Database)\s*:=\s*(.+);\s*$",
    re.IGNORECASE,
)
MIME_RE = re.compile(r"MimeDecodeString\s*\(\s*'([A-Za-z0-9+/=]+)'\s*\)", re.IGNORECASE)
ALT_HOST_COMMENT = re.compile(
    r"Base10\.Database\.Host\s*:=\s*'((?:\d{1,3}\.){3}\d{1,3})'",
    re.IGNORECASE,
)


def _strip_delphi_comments_preserve_strings(src: str) -> str:
    """{ ... } 블록 주석 제거 (문자열 리터럴 내부 ``{`` 는 거의 없음)."""
    out: list[str] = []
    i = 0
    n = len(src)
    in_squote = False
    while i < n:
        c = src[i]
        if c == "'" and (i == 0 or src[i - 1] != "\\"):
            in_squote = not in_squote
            out.append(c)
            i += 1
            continue
        if not in_squote and c == "{":
            depth = 1
            i += 1
            while i < n and depth:
                if src[i] == "{":
                    depth += 1
                elif src[i] == "}":
                    depth -= 1
                i += 1
            out.append("\n")
            continue
        out.append(c)
        i += 1
    return "".join(out)


def _decode_value(expr: str) -> str | None:
    expr = expr.strip()
    m = MIME_RE.search(expr)
    if m:
        try:
            return base64.b64decode(m.group(1)).decode("ascii", errors="replace")
        except Exception:
            return None
    m2 = re.search(r"'([^']*)'", expr)
    if m2:
        return m2.group(1)
    return None


def _line_without_cpp_comment(line: str) -> str:
    """``//`` 이후는 제거 (문자열 안 ``//`` 는 거의 없음)."""
    if "//" in line:
        return line.split("//", 1)[0].rstrip()
    return line.rstrip()


def _scan_chul_pas(path: Path) -> dict:
    raw = path.read_text(encoding="utf-8", errors="replace")
    stripped = _strip_delphi_comments_preserve_strings(raw)

    last: dict[str, str | None] = {
        "Host": None,
        "Login": None,
        "Password": None,
        "Database": None,
    }
    lines = stripped.splitlines()
    connected_idx: int | None = None
    for idx, line in enumerate(lines):
        if re.search(r"Base10\.Database\.Connected\s*:=\s*True", line, re.IGNORECASE):
            connected_idx = idx
            break

    segment = lines[:connected_idx] if connected_idx is not None else lines
    for line in segment:
        lc = _line_without_cpp_comment(line)
        if lc.strip().startswith("//"):
            continue
        m = ASSIGN_RE.match(lc)
        if not m:
            continue
        key_lower = m.group(1).lower()
        rhs = m.group(2)
        canon = {
            "host": "Host",
            "login": "Login",
            "password": "Password",
            "database": "Database",
        }.get(key_lower)
        if not canon:
            continue
        last[canon] = _decode_value(rhs)

    alt_hosts = sorted(set(ALT_HOST_COMMENT.findall(raw)))
    has_ini = bool(
        re.search(r"TIniFile\.Create\s*\(\s*GetExecPath\s*\+\s*'Config\.Ini'\s*\)", raw, re.IGNORECASE)
        or re.search(r"ReadString\s*\(\s*'Client'\s*,", raw, re.IGNORECASE)
    )
    has_connected = connected_idx is not None

    folder = path.parent.name
    rel = path.relative_to(WELOVE).as_posix()
    tree = rel.split("/", 1)[0] if "/" in rel else rel

    dbn = last.get("Database") or ""
    fam = ""
    if dbn.endswith("_db"):
        fam = dbn[: -len("_db")]
    elif "_" in dbn:
        fam = dbn.rsplit("_", 1)[0]

    mismatch = ""
    if fam and fam not in folder and not folder.startswith(fam):
        if "(" in folder:
            pref = folder.split("(", 1)[0]
            if pref != fam:
                mismatch = f"폴더명 접두({pref}) ≠ 활성 DB 패밀리({fam}) — 복붙/커스텀 빌드 의심"

    return {
        "tree": tree,
        "rel_path": rel,
        "build_folder": folder,
        "host": last.get("Host") or "",
        "database": last.get("Database") or "",
        "login": last.get("Login") or "",
        "password_pattern": last.get("Password") or "",
        "inferred_family": fam,
        "config_ini_client_reads": "예" if has_ini else "아니오",
        "alt_hosts_in_source_comments": ", ".join(alt_hosts) if alt_hosts else "",
        "folder_db_mismatch_note": mismatch,
        "connected_true_in_file": "예" if has_connected else "아니오",
    }


def _write_meta_sheet(wb: openpyxl.Workbook) -> None:
    sh = wb.create_sheet("조사 요약", 0)
    rows = [
        ("조사 대상", "WeLove_FTP/도서유통-New/**/Chul.pas, WeLove_FTP/도서유통-출판/**/Chul.pas"),
        ("결론(한 줄)", "총판·출판사별 DB/호스트/MySQL 로그인은 주로 Chul.pas 에 컴파일 시점 하드코딩(Base10.Database.*)."),
        ("보조 설정", "Config.Ini([Client] Base/Name/Uses/Pcip/Port) 읽기 코드가 있으나, 현재 본문의 DB 4종 대입은 대부분 Chul.pas 고정."),
        ("난독화", "MimeDecodeString(base64) 로 IP·DB명·*_user·*_pw 문자열을 인코딩한 빌드가 다수. 주석에 평문 예시가 병기된 경우 있음."),
        ("주의", "동일 폴더 계열이 아닌 중첩 경로(북앤북NEW/하위 타사 Chul.pas)는 '빌드 묶음' 단위로 별도 행 — 실제 배포 바이너리와 1:1 확인 필요."),
        ("스캔 한계", "일부 루트 Chul.pas 는 ``Base10.Database.Connected:=True`` 가 없어 파일 전체에서 마지막 대입만 취함 — 수동 검증 권장."),
        ("비밀 정책", "본 엑셀은 *_pw 명명 규칙만 기록. 운영 MySQL root 비밀번호는 servers.yaml / vault — 여기 미포함."),
        ("출력 경로", str(OUT_XLSX.relative_to(ROOT))),
    ]
    for ri, (k, v) in enumerate(rows, start=1):
        a = sh.cell(row=ri, column=1, value=k)
        a.font = Font(bold=True)
        b = sh.cell(row=ri, column=2, value=v)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    sh.column_dimensions["A"].width = 28
    sh.column_dimensions["B"].width = 100


def main() -> int:
    if not WELOVE.is_dir():
        print(f"[ERR] {WELOVE} 없음")
        return 2

    paths: list[Path] = []
    for pat in GLOBS:
        paths.extend(sorted(WELOVE.glob(pat)))

    rows_out: list[dict] = []
    for p in paths:
        try:
            rows_out.append(_scan_chul_pas(p))
        except Exception as e:
            rows_out.append(
                {
                    "tree": "",
                    "rel_path": str(p.relative_to(WELOVE)),
                    "build_folder": "",
                    "host": "",
                    "database": "",
                    "login": "",
                    "password_pattern": "",
                    "inferred_family": "",
                    "config_ini_client_reads": "",
                    "alt_hosts_in_source_comments": "",
                    "folder_db_mismatch_note": f"SCAN_ERR: {e}",
                    "connected_true_in_file": "",
                }
            )

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_meta_sheet(wb)
    sh = wb.create_sheet("Chul.pas 라우팅")
    headers = [
        "소스 트리",
        "상대 경로(WeLove_FTP 기준)",
        "빌드 폴더명",
        "활성 Host",
        "활성 Database",
        "활성 Login",
        "활성 Password(명명)",
        "DB명에서 추론한 account_family",
        "Config.Ini Client 읽기",
        "소스 내 다른 IP 후보(주석 등)",
        "폴더명 vs DB 불일치 메모",
        "Connected:=True 존재",
    ]
    for ci, h in enumerate(headers, start=1):
        c = sh.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for ri, r in enumerate(rows_out, start=2):
        sh.cell(row=ri, column=1, value=r["tree"])
        sh.cell(row=ri, column=2, value=r["rel_path"])
        sh.cell(row=ri, column=3, value=r["build_folder"])
        sh.cell(row=ri, column=4, value=r["host"])
        sh.cell(row=ri, column=5, value=r["database"])
        sh.cell(row=ri, column=6, value=r["login"])
        sh.cell(row=ri, column=7, value=r["password_pattern"])
        sh.cell(row=ri, column=8, value=r["inferred_family"])
        sh.cell(row=ri, column=9, value=r["config_ini_client_reads"])
        sh.cell(row=ri, column=10, value=r["alt_hosts_in_source_comments"])
        sh.cell(row=ri, column=11, value=r["folder_db_mismatch_note"])
        sh.cell(row=ri, column=12, value=r.get("connected_true_in_file", ""))
    for ci in range(1, len(headers) + 1):
        sh.column_dimensions[get_column_letter(ci)].width = 20 if ci == 12 else 22

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_XLSX))
    print(f"[OK] {len(rows_out)} files → {OUT_XLSX.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
