#!/usr/bin/env python3
"""ONE-OFF — ``DB정보 엑셀정리.xlsx`` 전 테넌트 DB 의 실 Id_Logn 행 덤프 (1회성, 사용 후 삭제).

목적
----
현재 웹 시스템(``도서물류관리프로그램/backend/app/services/auth_service.py``)이 실제로
검증하는 것은 ``Id_Logn.Gcode`` / ``Id_Logn.Gpass`` 두 컬럼이다. ``DB_FTP 엑셀정리.xlsx``
의 "ID/Pw" 는 셋업 시점 메모일 뿐이므로, 살아있는 정본을 직접 떠야 한다.

범위
----
- **기본**: MAN-017 엑셀의 **모든** 물류사(업체명) 행 → 총판·소속 출판사 DB·독립 출판사·
  테스트/기타 계정까지 ``Id_Logn`` 전부 (동일 ``db_name`` 중복 행은 1회만 쿼리).
- ``--distributors-only``: 예전 동작 — ``tools/seed_tenants_directory.py`` 의
  ``_PRIMARY_DIST_FAMILIES`` (총판 정본 9건) 만.

실 운영 토폴로지
----------------
- 서버1/2/3 은 외부 3306 직접 접속 불가 → SSH 터널 + (서버1/2 는 mysql3 raw protocol).
- 서버4 만 직접 연결.
- DB 사용자는 모든 서버 ``root`` 단일 — 테넌트 DB 명만 ``chul_05_db`` 등으로 USE.

따라서 본 스크립트는 자체적으로 SSH 터널을 구성하지 않고, **운영 백엔드의
`execute_query()` 를 in-process 로 호출**한다 (DRY — 가장 신뢰 가능한 경로).

전제
----
- ``도서물류관리프로그램/backend/servers.yaml`` 에 4개 서버 자격증명(SSH+MySQL) 채워짐.
- ``root`` 권한이 모든 테넌트 DB 에 cross-DB SELECT 가능 (운영 표준).

정책 정합 (docs/secrets-policy.md)
----------------------------------
- 출력은 ``WeLove_FTP/`` 하위 (gitignored) 에만 작성.
- 스크립트 자체에 자격증명 0줄.
- 사용 후 본 스크립트와 출력 xlsx 를 함께 ``rm -P`` 로 영구 삭제.

사용
----
    # 1) 어떤 서버/DB 를 덤프할지 미리 보기 (연결 안 함)
    python3 tools/_oneoff_dump_idlogn_for_distributors.py

    # 2) 실 덤프 (전 테넌트 — 독립 출판사·기타 포함)
    python3 tools/_oneoff_dump_idlogn_for_distributors.py --apply

    # 3) 총판 9건만 (예전 범위)
    python3 tools/_oneoff_dump_idlogn_for_distributors.py --apply --distributors-only

    # 4) 단일 서버만
    python3 tools/_oneoff_dump_idlogn_for_distributors.py --apply --server 서버3
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as _dt
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
BACKEND = ROOT / "도서물류관리프로그램" / "backend"
DB_INFO_XLSX = (
    ROOT
    / "WeLove_FTP/Welove_인수인계/셋팅방법/DB정보, DB_FTP 엑셀/DB정보 엑셀정리.xlsx"
)
OUT_PATH = ROOT / "WeLove_FTP" / "_oneoff_idlogn_dump_DELETE_AFTER_USE.xlsx"

# tools/seed_tenants_directory.py 와 동기 (ACTR / tenants 시드 추론)
_FAMILY_TYPE_MAP: dict[str, tuple[str, str]] = {
    "chul_09": ("T3", "warehouse_publisher"),
    "book_21": ("T3", "warehouse_publisher"),
    "book_07": ("T3", "warehouse_publisher"),
    "book_kb": ("T2_DIST", "distributor"),
    "kb_book": ("T2_DIST", "distributor"),
}
_DIST_FAMILIES = {f"chul_{i:02d}" for i in range(1, 9)} | {"chul_05", "chul_06", "chul_07", "chul_08"}
_PUB_FAMILIES = {f"book_{i:02d}" for i in range(1, 14)} | {
    "book_10", "book_11", "book_12", "book_13",
    "book_91", "book_99", "book_bs", "book_gs", "book_js", "sky_01",
}
_PRIMARY_DIST_FAMILIES = {f"chul_{i:02d}" for i in range(1, 9)} | {"book_kb", "kb_book"}

# 엑셀 라벨(서버1/2/3/4) → servers.yaml id 매핑 (servers.example.yaml 정본 기준)
_LABEL_TO_REMOTE_ID = {
    "서버1": "remote_154",
    "서버2": "remote_155",
    "서버3": "remote_153",
    "서버4": "remote_138",
}


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _infer_account_type(family: str) -> tuple[str, str]:
    """``tools/seed_tenants_directory.py::_infer_account_type`` 와 동기 (+ kb_book)."""
    if family in _FAMILY_TYPE_MAP:
        return _FAMILY_TYPE_MAP[family]
    if family in _DIST_FAMILIES or family.startswith("chul_"):
        return "T2_DIST", "distributor"
    if family in _PUB_FAMILIES or family.startswith("book_"):
        return "T3", "publisher"
    return "T3", "publisher"


def _segment_kr(account_type: str, build_role: str) -> str:
    if account_type == "T2_DIST" and build_role == "distributor":
        return "총판(물류)"
    if account_type.startswith("T3_WAREHOUSE") or build_role == "warehouse_publisher":
        return "물류내장형 출판(총판 DB 내 위러브 등)"
    if account_type == "T3" and build_role == "publisher":
        return "독립 출판사·일반 출판 DB(book_* 등)"
    return f"기타 ({account_type}/{build_role})"


def load_db_routes(*, only_distributors: bool) -> list[dict]:
    """`DB정보 엑셀정리.xlsx` 에서 라우트 로드. ``only_distributors`` 시 총판 정본만."""
    wb = openpyxl.load_workbook(str(DB_INFO_XLSX), data_only=True)
    sh = wb["Sheet1"]
    seen: set[tuple[str, str]] = set()
    rows: list[dict] = []
    for row in sh.iter_rows(values_only=True):
        cells = [_norm(v) for v in row]
        if len(cells) < 7:
            continue
        if "서버" in cells and "물류사명(업체명)" in cells:
            continue
        server_label, tenant, db_user, _db_pass, db_name = (
            cells[2], cells[3], cells[4], cells[5], cells[6],
        )
        if not server_label or not tenant or not db_user or not db_name:
            continue
        family = db_user[:-5] if db_user.endswith("_user") else db_user.split("_user", 1)[0]
        if only_distributors and family not in _PRIMARY_DIST_FAMILIES:
            continue
        sid_label = (
            server_label.split("(", 1)[1].rstrip(")").strip()
            if "(" in server_label else server_label.strip()
        )
        dedupe_key = (sid_label, db_name.strip().lower())
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        acc_type, build_role = _infer_account_type(family)
        rows.append(
            {
                "server_label": server_label,
                "server_id_label": sid_label,
                "remote_id": _LABEL_TO_REMOTE_ID.get(sid_label, ""),
                "tenant_name_kor": tenant,
                "account_family": family,
                "db_name": db_name.strip(),
                "account_type": acc_type,
                "build_role": build_role,
                "tenant_segment_kr": _segment_kr(acc_type, build_role),
            }
        )
    return rows


async def _dump_one(route: dict, sql: str) -> tuple[list[dict] | None, str | None]:
    """Run cross-DB SELECT via 운영 백엔드 execute_query."""
    from app.core.db import execute_query  # 지연 임포트
    try:
        rows = await execute_query(route["remote_id"], sql, ())
        return list(rows or []), None
    except Exception as e:
        return None, f"{type(e).__name__}: {str(e)[:160]}"


async def _dump_all(routes: list[dict]) -> list[dict]:
    """라우트별 cross-DB Id_Logn SELECT — root 권한 가정."""
    results: list[dict] = []
    for r in routes:
        if not r["remote_id"]:
            results.append({"route": r, "rows": None, "err": "remote_id 매핑 미정"})
            continue
        # cross-DB: `db_name`.Id_Logn 직접 지정 (서버 default DB 와 무관)
        sql = (
            f"SELECT Gcode, Gpass, Hname, Hcode, Gname "
            f"FROM `{r['db_name']}`.Id_Logn "
            f"WHERE Gcode IS NOT NULL AND TRIM(Gcode) <> ''"
        )
        rows, err = await _dump_one(r, sql)
        results.append({"route": r, "rows": rows, "err": err})
        tag = "✗" if err else "✓"
        n = len(rows) if rows is not None else 0
        print(f"  {tag} {r['server_id_label']:<6} | {r['account_family']:<8} | "
              f"{r['tenant_name_kor']:<14} | rows={n} | {err or ''}")
    return results


def write_warning_sheet(wb: openpyxl.Workbook) -> None:
    sh = wb.create_sheet("⚠ 사용 전 필독", 0)
    today = _dt.date.today().isoformat()
    delete_by = (_dt.date.today() + _dt.timedelta(days=7)).isoformat()
    msgs = [
        ("⚠️ 본 파일은 1회성 운반용입니다 — 사용 즉시 영구 삭제 ⚠️", True, "FFC7CE"),
        ("", False, None),
        (f"생성일: {today}", False, None),
        (f"삭제 권장일: {delete_by} (생성일 +7일 이내)", True, None),
        ("", False, None),
        ("정책 (docs/secrets-policy.md)", True, "BDD7EE"),
        ("- SEC-POL-CLASS-RED — 본 파일은 RED 등급 (Id_Logn 평문 비밀번호 포함).", False, None),
        ("- SEC-POL-STORAGE-01 — WeLove_FTP/ 는 .gitignore 로 전체 제외 위치.", False, None),
        ("- SEC-POL-STORAGE-04 — 외부 클라우드(Drive·Dropbox 등) 업로드 금지.", False, None),
        ("- 회의·PR·이슈·메신저 본문 평문 인용 금지 (마스킹 메타만 허용).", False, None),
        ("", False, None),
        ("출처", True, "BDD7EE"),
        ("- 라이브 쿼리: 운영 백엔드 execute_query() 를 in-process 호출.", False, None),
        ("- DB 자격증명: 도서물류관리프로그램/backend/servers.yaml (gitignored).", False, None),
        ("- SQL: SELECT Gcode, Gpass, Hname, Hcode, Gname FROM `<db>`.Id_Logn", False, None),
        ("- 분류 기준: tools/seed_tenants_directory.py 의 account_family → account_type/build_role 추론.", False, None),
        ("- '소속 출판사' 여부는 Id_Logn 만으로는 판별 불가 — 엑셀 물류사명·별도 G1/G7 계약 테이블과 대조.", False, None),
        ("", False, None),
        ("로그인 매핑 (docs/decision-rbac-and-id-logn-truth.md)", True, "BDD7EE"),
        ("- 로그인 ID 입력 = Id_Logn.Gcode (사번/Logn2)", False, None),
        ("- 비밀번호 입력 = Id_Logn.Gpass (평문)", False, None),
        ("- 창고/서버 선택 = 본 시트의 'warehouse_code 입력값' 컬럼 (=remote_id)", False, None),
        ("", False, None),
        ("사용 후", True, "FFC7CE"),
        ("- vault / 1Password 이관 후 본 xlsx + 본 스크립트 즉시 영구 삭제.", False, None),
        ("- git 추적 발견 시: 정책 §6 절차 — history rewrite + 자격증명 회전.", False, None),
    ]
    for ri, (text, bold, fill_hex) in enumerate(msgs, start=1):
        cell = sh.cell(row=ri, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            cell.font = Font(bold=True, size=12)
        if fill_hex:
            cell.fill = PatternFill("solid", fgColor=fill_hex)
    sh.column_dimensions["A"].width = 110


def write_summary_sheet(wb: openpyxl.Workbook, results: list[dict]) -> None:
    sh = wb.create_sheet("연결 결과 요약")
    headers = [
        "서버", "remote_id", "물류사명", "account_family", "db_name",
        "account_type", "build_role", "테넌트 구분(추론)",
        "Id_Logn 행수", "에러",
    ]
    for ci, h in enumerate(headers, start=1):
        c = sh.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for ri, r in enumerate(results, start=2):
        sh.cell(row=ri, column=1, value=r["route"]["server_id_label"])
        sh.cell(row=ri, column=2, value=r["route"]["remote_id"])
        sh.cell(row=ri, column=3, value=r["route"]["tenant_name_kor"])
        sh.cell(row=ri, column=4, value=r["route"]["account_family"])
        sh.cell(row=ri, column=5, value=r["route"]["db_name"])
        sh.cell(row=ri, column=6, value=r["route"]["account_type"])
        sh.cell(row=ri, column=7, value=r["route"]["build_role"])
        sh.cell(row=ri, column=8, value=r["route"]["tenant_segment_kr"])
        sh.cell(row=ri, column=9, value=len(r["rows"]) if r["rows"] is not None else 0)
        sh.cell(row=ri, column=10, value=r["err"] or "")
    for ci in range(1, len(headers) + 1):
        sh.column_dimensions[get_column_letter(ci)].width = 22


def write_idlogn_sheet(wb: openpyxl.Workbook, results: list[dict]) -> None:
    sh = wb.create_sheet("Id_Logn 덤프 (전체)")
    headers = [
        "서버(라벨)", "remote_id", "물류사명(DB)", "account_family",
        "account_type", "build_role", "테넌트 구분(추론)",
        "Gcode (=로그인 ID)", "Gpass (=비밀번호)",
        "Hname (조직/출판사명)", "Hcode (물류사코드)", "Gname (작업자/표시명)",
        "warehouse_code 입력값",
    ]
    for ci, h in enumerate(headers, start=1):
        c = sh.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    ri = 2
    for r in results:
        if not r["rows"]:
            continue
        for row in r["rows"]:
            sh.cell(row=ri, column=1, value=r["route"]["server_id_label"])
            sh.cell(row=ri, column=2, value=r["route"]["remote_id"])
            sh.cell(row=ri, column=3, value=r["route"]["tenant_name_kor"])
            sh.cell(row=ri, column=4, value=r["route"]["account_family"])
            sh.cell(row=ri, column=5, value=r["route"]["account_type"])
            sh.cell(row=ri, column=6, value=r["route"]["build_role"])
            sh.cell(row=ri, column=7, value=r["route"]["tenant_segment_kr"])
            sh.cell(row=ri, column=8, value=_norm(row.get("Gcode")))
            sh.cell(row=ri, column=9, value=_norm(row.get("Gpass")))
            sh.cell(row=ri, column=10, value=_norm(row.get("Hname")))
            sh.cell(row=ri, column=11, value=_norm(row.get("Hcode")))
            sh.cell(row=ri, column=12, value=_norm(row.get("Gname")))
            sh.cell(row=ri, column=13, value=r["route"]["remote_id"])
            ri += 1
    for ci in range(1, len(headers) + 1):
        sh.column_dimensions[get_column_letter(ci)].width = 22


def main() -> int:
    ap = argparse.ArgumentParser(description="Id_Logn 1회성 덤프 (MAN-017 전 테넌트)")
    ap.add_argument("--apply", action="store_true",
                    help="실제 백엔드 execute_query 호출 (생략 시 dry-run).")
    ap.add_argument(
        "--distributors-only",
        action="store_true",
        help="총판 정본 9건만 (chul_01~08, book_kb, kb_book).",
    )
    ap.add_argument("--server", default="",
                    help="서버 라벨 한정 (예: 서버2). 미지정 시 전체.")
    ap.add_argument("--family", default="",
                    help="account_family 한정 (예: chul_01). 미지정 시 전체.")
    args = ap.parse_args()

    routes = load_db_routes(only_distributors=bool(args.distributors_only))
    if args.server:
        routes = [r for r in routes if r["server_id_label"] == args.server]
    if args.family:
        routes = [r for r in routes if r["account_family"] == args.family]

    scope = "총판만" if args.distributors_only else "전 테넌트(MAN-017)"
    print(f"[{scope}] 라우트 {len(routes)} 건 (db_name 기준 중복 제거)")
    for r in routes:
        rid = r["remote_id"] or "(미매핑)"
        print(f"  - {r['server_id_label']:<6} | {rid:<11} | "
              f"{r['account_family']:<8} | {r['tenant_name_kor']:<14} | db={r['db_name']}")

    if not args.apply:
        print()
        print("[DRY-RUN] 실 호출 없음.  실 덤프: --apply")
        return 0

    print()
    print(f"[APPLY] 운영 백엔드 execute_query() 통해 {len(routes)} 건 덤프 시작")

    # backend in-process 임포트 준비
    sys.path.insert(0, str(BACKEND))
    os.chdir(str(BACKEND))  # servers.yaml 상대경로 해석을 위해

    results = asyncio.run(_dump_all(routes))

    # 결과 파일 작성 (gitignored)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    write_warning_sheet(wb)
    write_summary_sheet(wb, results)
    write_idlogn_sheet(wb, results)
    wb.save(str(OUT_PATH))

    succ = sum(1 for r in results if r["err"] is None)
    fail = sum(1 for r in results if r["err"])
    total_rows = sum(len(r["rows"]) for r in results if r["rows"])
    print()
    print(f"[OK] 출력: {OUT_PATH.relative_to(ROOT)}  (gitignored)")
    print(f"     성공 {succ} / 실패 {fail} / 총 Id_Logn rows = {total_rows}")
    print(f"[중요] 사용 즉시 본 xlsx + 본 스크립트를 영구 삭제하십시오.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
