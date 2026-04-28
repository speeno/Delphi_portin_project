#!/usr/bin/env python3
"""기상청 기상특보 구역 코드 자료를 경량 JSON으로 변환한다."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Iterable


def _rows_from_csv(path: Path) -> Iterable[dict[str, object]]:
    for enc in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            f = path.open("r", encoding=enc, newline="")
            sample = f.read(1024)
            f.seek(0)
            if sample:
                with f:
                    yield from csv.DictReader(f)
                return
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        yield from csv.DictReader(f)


def _rows_from_xlsx(path: Path) -> Iterable[dict[str, object]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        raise SystemExit("xlsx 입력은 openpyxl 설치가 필요합니다. CSV로 저장 후 다시 실행하세요.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(x or "").strip() for x in next(rows)]
    for row in rows:
        yield {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}


def _pick(row: dict[str, object], *keys: str) -> str:
    compact = {str(k).lower().replace(" ", "").replace("_", ""): v for k, v in row.items()}
    for key in keys:
        norm = key.lower().replace(" ", "").replace("_", "")
        if norm in compact:
            return str(compact[norm] or "").strip()
    return ""


def _as_float(raw: str) -> float | None:
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _as_int(raw: str) -> int | None:
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return None


def _normalize(row: dict[str, object]) -> dict[str, object] | None:
    area_code = _pick(row, "특보구역코드", "구역코드", "areaCode", "warnAreaCd", "code")
    name = _pick(row, "특보구역명", "구역명", "지역명", "name", "areaName")
    stn_id = _pick(row, "지점번호", "지점코드", "stnId", "stn")
    nx = _as_int(_pick(row, "nx", "격자 X", "격자X", "X"))
    ny = _as_int(_pick(row, "ny", "격자 Y", "격자Y", "Y"))
    lat = _as_float(_pick(row, "위도", "lat", "latitude"))
    lon = _as_float(_pick(row, "경도", "lon", "longitude"))
    if not area_code and not stn_id:
        return None
    return {
        "area_code": area_code,
        "name": name or area_code or stn_id,
        "stn_id": stn_id,
        "lat": lat,
        "lon": lon,
        "nx": nx,
        "ny": ny,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("input_path")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    source = Path(args.input_path)
    rows = _rows_from_xlsx(source) if source.suffix.lower() == ".xlsx" else _rows_from_csv(source)
    out = [norm for row in rows if (norm := _normalize(row))]
    target = Path(args.output)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {len(out)} KMA warning areas to {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
