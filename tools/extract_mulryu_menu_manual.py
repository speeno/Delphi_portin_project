#!/usr/bin/env python3
"""Extract WeLove 물류프로그램 메뉴얼.xlsx into MAN-009 documentation set.

Source: WeLove_FTP/Welove_인수인계/위러브 인수인계/물류프로그램 메뉴얼.xlsx (MAN-009)

Outputs:
  analysis/mulryu_program_menu_manual.json  — machine readable rows
  docs/welove-mulryu-program-menu-manual.md — human readable summary

The workbook is a per-sheet (top-level menu) operational manual. Each sheet
holds free-form rows: title rows, sub-menu headings, and free-text instructions.
We preserve the row-by-row content and group by sheet, exposing both the raw
rows (for traceability) and a curated submenu/feature list aggregated by simple
heading detection (rows where the first cell is short and the rest are blank).
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "WeLove_FTP/Welove_인수인계/위러브 인수인계/물류프로그램 메뉴얼.xlsx"
JSON_OUT = ROOT / "analysis/mulryu_program_menu_manual.json"
MD_OUT = ROOT / "docs/welove-mulryu-program-menu-manual.md"


def _norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def parse_sheet(sheet) -> dict:
    rows: list[list[str]] = []
    for r in range(1, sheet.max_row + 1):
        cells = [_norm(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        if not any(cells):
            rows.append([])  # blank separator
            continue
        rows.append(cells)

    submenus: list[dict] = []
    current_heading: str | None = None
    current_lines: list[str] = []

    def flush() -> None:
        nonlocal current_heading, current_lines
        if current_heading is not None:
            submenus.append({
                "heading": current_heading,
                "lines": [ln for ln in current_lines if ln],
            })
        current_heading = None
        current_lines = []

    heading_re = re.compile(r"^[\d\s\.\-가-힣A-Za-z()/]+$")
    for cells in rows:
        if not cells:
            continue
        first = cells[0]
        rest = [c for c in cells[1:] if c]
        if first and not rest and 1 <= len(first) <= 30 and heading_re.match(first):
            flush()
            current_heading = first
            continue
        line = " | ".join(c for c in cells if c)
        if line:
            current_lines.append(line)

    flush()

    return {
        "name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "rows": rows,
        "submenus": submenus,
    }


def render_md(payload: dict) -> str:
    lines: list[str] = []
    lines.append("# 위러브 물류프로그램 메뉴얼 (MAN-009)")
    lines.append("")
    lines.append("| 항목 | 내용 |")
    lines.append("|------|------|")
    lines.append("| 작성일 | 2026-04-24 |")
    lines.append(
        "| 원천 | "
        "[`WeLove_FTP/Welove_인수인계/위러브 인수인계/물류프로그램 메뉴얼.xlsx`]"
        "(../WeLove_FTP/Welove_인수인계/위러브 인수인계/물류프로그램 메뉴얼.xlsx) (`MAN-009`) |"
    )
    lines.append("| 추출 도구 | [`tools/extract_mulryu_menu_manual.py`](../tools/extract_mulryu_menu_manual.py) |")
    lines.append(
        "| JSON 정본 | "
        "[`analysis/mulryu_program_menu_manual.json`](../analysis/mulryu_program_menu_manual.json) |"
    )
    lines.append("| 추적 ID | 시트 단위 `MAN-009-<SHEET>`, 서브메뉴 단위 `MAN-009-<SHEET>.<HEADING>` |")
    lines.append("")
    lines.append("> 본 문서는 레거시 위러브 물류 프로그램 운영 매뉴얼의 **시트별 행 기반 추출본**입니다.")
    lines.append("> 추출 시점에 행 그룹핑 휴리스틱을 적용했지만, **세부 의미는 원본 xlsx 가 우선**입니다.")
    lines.append("")
    lines.append("## 시트 카탈로그")
    lines.append("")
    lines.append("| 시트 | 행 | 열 | 추출된 서브메뉴 수 |")
    lines.append("|------|----|----|--------------------|")
    for s in payload["sheets"]:
        lines.append(
            f"| `{s['name']}` | {s['max_row']} | {s['max_column']} | {len(s['submenus'])} |"
        )
    lines.append("")
    for s in payload["sheets"]:
        lines.append(f"## `{s['name']}` — 서브메뉴")
        lines.append("")
        if not s["submenus"]:
            lines.append("(추출된 서브메뉴 헤딩 없음 — 원본 xlsx 직접 열람 필요)")
            lines.append("")
            continue
        for sub in s["submenus"]:
            lines.append(f"### `{sub['heading']}`")
            lines.append("")
            for ln in sub["lines"][:50]:
                lines.append(f"- {ln}")
            if len(sub["lines"]) > 50:
                lines.append(f"- … (총 {len(sub['lines'])} 줄, 50줄 표시)")
            lines.append("")
    return "\n".join(lines) + "\n"


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"원본 파일 없음: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    sheets = [parse_sheet(s) for s in wb.worksheets]
    payload = {
        "source": str(XLSX.relative_to(ROOT)),
        "extracted_by": "tools/extract_mulryu_menu_manual.py",
        "tracking_id": "MAN-009",
        "sheets": sheets,
    }
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    MD_OUT.write_text(render_md(payload), encoding="utf-8")
    print(f"wrote {JSON_OUT.relative_to(ROOT)}")
    print(f"wrote {MD_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
