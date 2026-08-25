"""DEC-198 — 엑셀 내보내기 = 화면에 보이는 컬럼·순서 / 도서별판매 도서분류 응답 누락.

사용자 원칙(2026-08-25): "엑셀 출력은 화면에 보이도록 설정된 필드가 동일한 순서로 출력되어야
한다." + 리포트: "도서별판매 엑셀 다운 시 누락 출력 컬럼이 있다", "도서별판매 화면에 도서
분류가 출력되지 않는다".

원인
----
1. 엑셀은 서버 고정 목록(`_BOOK_SALES_EXPORT_COLUMNS`)이라 화면의 도서분류·판매수량·
   판매금액·재고 3종이 빠졌고, 사용자가 컬럼 설정으로 바꾼 순서/숨김도 반영되지 않았다.
2. `BookSalesRow` 응답 모델에 `sname`/`gubun_code` 가 없어 서비스가 부착한 도서분류를
   FastAPI response_model 이 잘라냈다(DEC-169 ISBN 누락과 같은 유형).

사용자 규칙: test 폴더에 저장.
"""

from __future__ import annotations

import io
import json
import sys
from pathlib import Path
from unittest import TestCase, main
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "도서물류관리프로그램" / "backend"
FRONT = ROOT / "도서물류관리프로그램" / "frontend" / "src"
sys.path.insert(0, str(BACKEND))

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.main import app  # noqa: E402
from app.models.inquiry import BookSalesRow  # noqa: E402
from app.routers.auth import get_current_user  # noqa: E402
from app.services import reports_service  # noqa: E402


def _auth() -> dict:
    return {"user_id": "hong01", "server_id": "remote_1", "hcode": "5019"}


app.dependency_overrides[get_current_user] = _auth
Q = "?serverId=remote_1&dateFrom=2026-08-01&dateTo=2026-08-24"


def _book_rows():
    return [{
        "gcode": "00100", "gname": "피복인간공학*", "gdang": 14450, "gisbn": "978-89-363-0001-1",
        "gdate": "2026.08.10", "giqut": 0, "goqut": 3, "gjqut": 0, "gbqut": -1, "gpqut": 0,
        "gosum": 43350, "gbsum": -14450, "gpsum": 0, "gisum": 0,
        "sname": "의류, 의상학", "gubun_code": "10003",
    }]


def _sheet(res) -> tuple[list, list]:
    wb = load_workbook(io.BytesIO(res.content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[0]), [list(r) for r in rows[1:]]


class BookSalesExportTests(TestCase):
    def setUp(self) -> None:
        app.dependency_overrides[get_current_user] = _auth
        self.client = TestClient(app)

    def _export(self, columns):
        async def fake(**kw):
            return {"rows": _book_rows(), "total": 1,
                    "page": {"limit": 500, "offset": 0, "total": 1, "has_more": False}}

        async def fake_stock(server_id, rows, *, hcode, asof):
            for r in rows:
                r["hq_stock"], r["wh_stock"], r["total_stock"] = 7, 8, 15

        with patch.object(reports_service, "get_book_sales", side_effect=fake), \
             patch.object(reports_service, "attach_period_end_stock", side_effect=fake_stock):
            q = Q + (f"&columns={json.dumps(columns, ensure_ascii=False)}" if columns is not None else "")
            return self.client.get("/api/v1/reports/book-sales/export.xlsx" + q)

    def test_headers_follow_visible_columns_in_order(self) -> None:
        """화면이 넘긴 키·라벨·순서 그대로 — 파생 컬럼(판매수량/판매금액/재고)도 값이 찬다."""
        cols = [
            {"key": "sname", "label": "도서분류"}, {"key": "gname", "label": "도서명"},
            {"key": "sellQut", "label": "판매수량"}, {"key": "sellSum", "label": "판매금액"},
            {"key": "totalStock", "label": "재고합계"}, {"key": "gcode", "label": "도서코드"},
        ]
        res = self._export(cols)
        self.assertEqual(res.status_code, 200, res.text[:200])
        headers, rows = _sheet(res)
        self.assertEqual(headers, ["도서분류", "도서명", "판매수량", "판매금액", "재고합계", "도서코드"])
        self.assertEqual(rows[0], ["의류, 의상학", "피복인간공학*", 2, 28900, 15, "00100"])

    def test_hidden_columns_are_not_exported(self) -> None:
        res = self._export([{"key": "gname", "label": "도서명"}])
        headers, _ = _sheet(res)
        self.assertEqual(headers, ["도서명"])

    def test_unknown_key_rejected(self) -> None:
        res = self._export([{"key": "password", "label": "x"}])
        self.assertEqual(res.status_code, 422)

    def test_omitted_columns_keeps_default_list(self) -> None:
        """호환 — columns 미전달(구 클라이언트)은 종전 기본 목록."""
        res = self._export(None)
        headers, _ = _sheet(res)
        self.assertEqual(headers[:2], ["코드", "도서명"])


class CustomerSalesExportTests(TestCase):
    def setUp(self) -> None:
        app.dependency_overrides[get_current_user] = _auth
        self.client = TestClient(app)

    def test_headers_follow_visible_columns(self) -> None:
        async def fake(**kw):
            return {"rows": [{"hcode": "5019", "gcode": "00001", "gname": "(주)교보문고", "gjisa": "",
                              "gdate": "2026.08.20", "goqut": 488, "gosum": 14411450, "gjqut": 0,
                              "gbqut": 0, "gbsum": 0, "gsusu": 488, "gjsum": 85629320, "gssum": 14411450,
                              "cust_gubun": "인터넷서점", "cust_jubun": "서 울"}],
                    "total": 1, "page": {"limit": 500, "offset": 0, "total": 1, "has_more": False}}

        cols = [{"key": "cust_gubun", "label": "거래처구분"}, {"key": "gname", "label": "거래처명"},
                {"key": "gjsum", "label": "수금액"}, {"key": "gssum", "label": "판매금액"}]
        with patch.object(reports_service, "get_customer_sales", side_effect=fake):
            res = self.client.get("/api/v1/reports/customer-sales/export.xlsx" + Q
                                  + "&columns=" + json.dumps(cols, ensure_ascii=False))
        self.assertEqual(res.status_code, 200, res.text[:200])
        headers, rows = _sheet(res)
        self.assertEqual(headers, ["거래처구분", "거래처명", "수금액", "판매금액"])
        self.assertEqual(rows[0], ["인터넷서점", "(주)교보문고", 85629320, 14411450])


class BookClassResponseTests(TestCase):
    def test_model_keeps_sname(self) -> None:
        """도서분류 — 모델에 필드가 없으면 response_model 이 잘라낸다."""
        self.assertIn("sname", BookSalesRow.model_fields)
        self.assertIn("gubun_code", BookSalesRow.model_fields)
        row = BookSalesRow.model_validate(_book_rows()[0]).model_dump()
        self.assertEqual(row["sname"], "의류, 의상학")


class ScreenWiringTests(TestCase):
    def test_pages_pass_visible_columns(self) -> None:
        for page in ("book-sales", "customer-sales"):
            src = (FRONT / "app" / "(app)" / "reports" / page / "page.tsx").read_text(encoding="utf-8")
            self.assertIn("columns: visibleColumns.map((c) => ({ key: c.id ?? c.key, label: c.label }))", src, page)


if __name__ == "__main__":
    main()
