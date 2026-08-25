"""DEC-198 일괄 적용 — 통계/반품/정산 목록 내보내기 전부 «화면에 보이는 컬럼·순서».

사용자(2026-08-25): "일괄 적용해 주세요." 대상 = 엑셀 저장 버튼이 있는 목록 화면 15곳 /
내보내기 라우트 12곳(도서별·거래처별판매 포함 시 14). 기초관리 마스터 내보내기(재입력
템플릿, 별도 `fields` 파라미터)는 제외.

공용 리졸버 `app.services.export_columns.resolve_export_columns` 규칙:
  fields(라우트 별칭/파생) → 기본 목록 키 → 행에 실제 있는 키 → 그 외 422.

사용자 규칙: test 폴더에 저장.
"""

from __future__ import annotations

import io
import json
import re
import sys
from pathlib import Path
from unittest import TestCase, main
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "도서물류관리프로그램" / "backend"
FRONT = ROOT / "도서물류관리프로그램" / "frontend" / "src"
sys.path.insert(0, str(BACKEND))

from fastapi import HTTPException  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.main import app  # noqa: E402
from app.routers.auth import get_current_user  # noqa: E402
from app.services.export_columns import resolve_export_columns  # noqa: E402


def _auth() -> dict:
    return {"user_id": "hong01", "server_id": "remote_1", "hcode": "5019", "role": "admin",
            "permissions": ["*"]}


app.dependency_overrides[get_current_user] = _auth


def _sheets(res) -> dict[str, tuple[list, list]]:
    wb = load_workbook(io.BytesIO(res.content), data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        out[ws.title] = (list(rows[0]) if rows else [], [list(r) for r in rows[1:]])
    return out


class ResolverTests(TestCase):
    DEFAULT = [("코드", "gcode"), ("이름", "gname"), ("라벨", lambda r: "L")]

    def test_none_spec_returns_default(self) -> None:
        self.assertEqual(resolve_export_columns(None, default=self.DEFAULT), self.DEFAULT)

    def test_order_and_labels_follow_spec(self) -> None:
        spec = json.dumps([{"key": "gname", "label": "이름!"}, {"key": "gcode", "label": "코드!"}])
        cols = resolve_export_columns(spec, default=self.DEFAULT)
        self.assertEqual(cols, [("이름!", "gname"), ("코드!", "gcode")])

    def test_fields_alias_and_row_presence(self) -> None:
        spec = json.dumps([{"key": "book_name", "label": "도서명"}, {"key": "extra", "label": "행키"}])
        cols = resolve_export_columns(
            spec, default=self.DEFAULT, fields={"book_name": "bname"}, rows=[{"extra": 1, "bname": "x"}],
        )
        self.assertEqual(cols, [("도서명", "bname"), ("행키", "extra")])

    def test_unknown_key_422(self) -> None:
        spec = json.dumps([{"key": "password", "label": "x"}])
        with self.assertRaises(HTTPException) as cm:
            resolve_export_columns(spec, default=self.DEFAULT, rows=[{"gcode": "1"}])
        self.assertEqual(cm.exception.status_code, 422)

    def test_bad_json_422(self) -> None:
        with self.assertRaises(HTTPException):
            resolve_export_columns("not-json", default=self.DEFAULT)


class RouteTests(TestCase):
    def setUp(self) -> None:
        app.dependency_overrides[get_current_user] = _auth
        self.client = TestClient(app)

    def test_sales_period_group_by_uses_label(self) -> None:
        from app.services import stats_service

        async def fake(**kw):
            return {"items": [{"bucket": "2026-08", "group_by": "monthly", "buy_qut_total": 1,
                               "buy_sum_total": 10, "qut_total": 2, "sum_total": 20}],
                    "totals": {"buy_qut_total": 1, "buy_sum_total": 10, "qut_total": 2, "sum_total": 20}}

        cols = [{"key": "sum_total", "label": "매출금액"}, {"key": "group_by", "label": "집계단위"}, {"key": "bucket", "label": "구간"}]
        with patch.object(stats_service, "get_sales_period", side_effect=fake):
            res = self.client.get("/api/v1/stats/sales-period/export.xlsx?serverId=remote_1&dateFrom=2026-08-01&dateTo=2026-08-24&groupBy=monthly&columns="
                                  + json.dumps(cols, ensure_ascii=False))
        self.assertEqual(res.status_code, 200, res.text[:300])
        (headers, rows), = _sheets(res).values()
        self.assertEqual(headers, ["매출금액", "집계단위", "구간"])
        self.assertEqual(rows[0][2], "2026-08")
        self.assertNotEqual(rows[0][1], "monthly", "group_by 는 화면처럼 표시 라벨로")

    def test_quarterly_month_sheet_follows_columns(self) -> None:
        from app.services import stats_service

        async def fake(**kw):
            return {"comparison": [{"label": "2026 Q3", "gsumx": 1, "gsumy": 2, "gssum": 3, "profit": 4}],
                    "items": [{"gdate": "2026-07", "gsumx": 10, "gsumy": 20, "gssum": 30}],
                    "totals": {"gsumx": 10, "gsumy": 20, "gssum": 30, "profit": 0}}

        cols = [{"key": "gssum", "label": "잔액"}, {"key": "gdate", "label": "월"}]
        with patch.object(stats_service, "get_quarterly_summary", side_effect=fake):
            res = self.client.get("/api/v1/stats/quarterly-summary/export.xlsx?serverId=remote_1&year=2026&quarter=3&columns="
                                  + json.dumps(cols, ensure_ascii=False))
        self.assertEqual(res.status_code, 200, res.text[:300])
        sheets = _sheets(res)
        self.assertEqual(sheets["월별 상세"][0], ["잔액", "월"])
        self.assertEqual(sheets["분기 비교"][0][0], "분기", "분기 비교 시트는 화면 고정 표 — 기본 유지")

    def test_returns_daily_follows_columns(self) -> None:
        from app.services import returns_service

        async def fake(**kw):
            return {"master": [{"gdate": "2026.08.01", "hcode": "5019", "hname": "교문사",
                                "line_count": 3, "total_qty": 5, "total_amount": 50000}]}

        cols = [{"key": "hname", "label": "출판사"}, {"key": "total_amount", "label": "금액"}]
        with patch.object(returns_service, "daily_report", side_effect=fake):
            res = self.client.get("/api/v1/returns/reports/daily/export.xlsx?serverId=remote_1&date_from=2026-08-01&date_to=2026-08-24&columns="
                                  + json.dumps(cols, ensure_ascii=False))
        self.assertEqual(res.status_code, 200, res.text[:300])
        (headers, rows), = _sheets(res).values()
        self.assertEqual(headers, ["출판사", "금액"])
        self.assertEqual(rows[0], ["교문사", 50000])


class ScreenWiringTests(TestCase):
    PAGES = """reports/book-sales reports/customer-sales reports/year-end-book
        returns/ledger returns/period-report returns/reports
        settlement/period settlement/shipping-ledger settlement/shipping-status
        stats/book-turnover stats/book stats/customer-analysis stats/customer stats/monthly
        stats/publisher stats/quarterly-summary stats/sales-period""".split()

    def test_every_export_page_passes_visible_columns(self) -> None:
        for pg in self.PAGES:
            with self.subTest(page=pg):
                src = (FRONT / "app" / "(app)" / pg / "page.tsx").read_text(encoding="utf-8")
                self.assertIn("columns: visibleColumns.map((c) => ({ key: c.id ?? c.key, label: c.label }))", src)

    def test_every_export_api_forwards_columns(self) -> None:
        for lib in ("inquiry-api", "stats-api", "returns-api", "settlement-api"):
            src = (FRONT / "lib" / f"{lib}.ts").read_text(encoding="utf-8")
            n_fn = len(re.findall(r"\w+ExportBlob: \(", src))
            n_cols = src.count("JSON.stringify(params.columns)") + src.count("JSON.stringify(p.columns)")
            # cashStatusExportBlob(settlement) 은 호출 화면이 없어 제외 대상.
            expected = n_fn - (1 if lib == "settlement-api" else 0)
            self.assertGreaterEqual(n_cols, expected, f"{lib}: ExportBlob {n_fn} vs columns {n_cols}")


if __name__ == "__main__":
    main()
