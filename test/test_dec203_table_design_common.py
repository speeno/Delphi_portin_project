"""DEC-203 — 표(목록) 공통 디자인 + 섹션 헤더 + 「내용 전체 보기」/엑셀/출력 (2026-08-25 14:30 결과 목업).

목업(도서별 수불원장 결과 화면·거래처 원장 디폴트)에서 추출해 **모든 화면 공통**으로 반영한 요소
--------------------------------------------------------------------------
- 표: 카드 프레임(둥근 모서리·테두리·그림자) 없음, 회색 헤더행(`table-head`)·진한 글자, 가로 구분선만,
  선택 행 민트(`row-selected`), 표 포커스 시 파란 테두리(`row-focus`), 합계행 회색·굵게.
- 표 위 섹션 헤더(`SectionHeader`): 굵은 제목 + 회색 메타, 오른쪽 액션.
- 「내용 전체 보기」: 체크하면 상단 표가 스크롤 없이 모두 보이도록 영역이 자동으로 커진다(분할 off).
- 「엑셀 다운로드」: 화면 표를 표시 컬럼 순서 그대로 범용 xlsx 라우트로(`POST /api/v1/export/table-xlsx`).
- 「출력」: 표만 새 창에 그려 브라우저 인쇄.
"""

from __future__ import annotations

import re
import sys
from io import BytesIO
from pathlib import Path
from unittest import IsolatedAsyncioTestCase, TestCase, main

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "도서물류관리프로그램" / "backend"))
FRONT = ROOT / "도서물류관리프로그램" / "frontend" / "src"
APP = FRONT / "app" / "(app)"


def _read(rel: str) -> str:
    return (FRONT / rel).read_text(encoding="utf-8")


class CommonTableStyle(TestCase):
    def test_tokens_defined_for_light_and_dark(self) -> None:
        css = _read("app/globals.css")
        root = css[: css.index(".dark {")]
        dark = css[css.index(".dark {") :]
        for tok in ("--table-head", "--row-selected", "--row-focus"):
            # 라이트는 :root 에 1회, 다크는 .dark 에 1회 — 다크 값이 :root 에 끼면 라이트 헤더가 어두워진다
            # (2026-08-25 삽입 오프셋 실수로 실제 발생).
            self.assertEqual(root.count(f"  {tok}:"), 1, f"{tok} :root")
            self.assertEqual(dark.count(f"  {tok}:"), 1, f"{tok} .dark")
            self.assertIn(f"--color-{tok[2:]}: var({tok});", css)
        self.assertIn("--table-head: oklch(0.94 0 0)", root)

    def test_list_table_constants_frameless(self) -> None:
        src = _read("components/data-grid/list-table-card.tsx")
        self.assertIn('"w-full min-w-0 overflow-x-auto bg-card"', src)
        self.assertNotIn("overflow-x-auto rounded-2xl border border-border bg-card shadow-sm", src)
        self.assertIn('LIST_TABLE_HEAD_CLASS = "bg-table-head"', src)
        self.assertIn("text-sm font-semibold tracking-[-0.02em] text-foreground", src)
        self.assertIn('LIST_TABLE_ROW_SELECTED_CLASS = "bg-row-selected"', src)
        self.assertIn("outline-row-focus", src)
        self.assertIn('"border-t border-border bg-muted font-semibold text-foreground"', src)

    def test_data_grid_uses_common_style(self) -> None:
        src = _read("components/data-grid/data-grid.tsx")
        self.assertIn("sticky top-0 z-10 bg-table-head px-4 py-3 text-sm font-semibold", src)
        self.assertIn("LIST_TABLE_ROW_SELECTED_CLASS} ${LIST_TABLE_ROW_FOCUS_CLASS}", src)
        self.assertNotIn('"bg-primary/10 ring-1 ring-primary/40"', src)
        self.assertIn("<tr className={LIST_TABLE_FOOTER_ROW_CLASS}>", src)
        self.assertIn('"group/dg ', src, "포커스 행 테두리는 group-focus-within/dg 로 켜진다")

    def test_no_inline_table_frame_left(self) -> None:
        pat = re.compile(r"overflow-(?:auto|x-auto|y-auto) rounded-2xl border border-border bg-card shadow-sm")
        left = [
            str(p.relative_to(FRONT))
            for p in list(APP.glob("**/*.tsx")) + list((FRONT / "components").glob("**/*.tsx"))
            if pat.search(p.read_text(encoding="utf-8"))
        ]
        self.assertEqual(left, [], left)

    def test_section_header_component(self) -> None:
        src = _read("components/shared/section-header.tsx")
        self.assertIn('data-slot="section-header"', src)
        self.assertIn("text-lg font-bold tracking-tight", src)
        self.assertIn("SECTION_ACTION_BUTTON_CLASS", src)


class LedgerScreens(TestCase):
    """대표 화면 2곳 — 도서별수불원장(/inventory/ledger)·거래처원장(/ledger/customer)."""

    def _check(self, rel: str, storage_key: str, hint: str, sel_var: str) -> None:
        src = _read(rel)
        self.assertIn(f"<EmptyHint>{hint}</EmptyHint>", src)
        self.assertIn(f'storageKey="{storage_key}"', src)
        self.assertIn(f"disabled={{{sel_var} === null || showAll}}", src, "전체 보기면 분할 off")
        self.assertIn('data-legacy-id="ShowAll"', src)
        self.assertIn("내용 전체 보기", src)
        # 전체 보기 = 스크롤 없이 펼침 / 아니면 분할 칸을 채우고 내부 스크롤
        self.assertIn('showAll ? "overflow-visible" : "min-h-0 flex-1 overflow-auto max-h-[calc(100dvh-14rem)]"', src)
        self.assertEqual(src.count("<SectionHeader"), 2)
        self.assertIn("엑셀 다운로드", src)
        self.assertIn("exportTableXlsx", src)
        self.assertIn("printTable", src)
        self.assertIn("LIST_TABLE_ROW_SELECTED_CLASS", src)
        self.assertIn("<tr className={LIST_TABLE_FOOTER_STICKY_CLASS}>", src)
        self.assertNotIn("bg-primary/10", src)
        self.assertNotIn("bg-primary/5", src)

    def test_book_ledger(self) -> None:
        self._check("app/(app)/inventory/ledger/page.tsx", "inventory.ledger", "거래일자와 도서명으로 검색하세요", "selDate")

    def test_customer_ledger(self) -> None:
        self._check(
            "app/(app)/ledger/customer/page.tsx", "ledger.customer", "거래일자와 거래처명/거래처 코드로 검색하세요", "selKey"
        )

    def test_export_columns_follow_visible_order(self) -> None:
        """엑셀 컬럼 = 화면 표 헤더 순서 (2026-08-25 사용자 규칙)."""
        src = _read("app/(app)/inventory/ledger/page.tsx")
        i = src.index("const TOP_COLUMNS = [")
        block = src[i : src.index("];", i)]
        labels = re.findall(r'label: "([^"]+)"', block)
        self.assertEqual(labels, ["거래일자", "입고", "반입", "출고", "증정", "반품", "폐기", "변경", "현재고", "재고(반)"])


class TableExportHelpers(TestCase):
    def test_frontend_helper_and_post_blob(self) -> None:
        helper = _read("lib/table-export.ts")
        self.assertIn('api.postBlob("/api/v1/export/table-xlsx"', helper)
        self.assertIn("export function printTable", helper)
        self.assertIn("window.print()", helper)
        client = _read("lib/api-client.ts")
        self.assertIn("postBlob: async (path: string, body: unknown", client)

    def test_backend_route_registered(self) -> None:
        main_py = (ROOT / "도서물류관리프로그램" / "backend" / "app" / "main.py").read_text(encoding="utf-8")
        self.assertIn("app.include_router(export_table.router)", main_py)


class TableExportRoute(IsolatedAsyncioTestCase):
    async def test_builds_xlsx_in_column_order(self) -> None:
        from openpyxl import load_workbook

        from app.routers.export_table import TableExportColumn, TableExportRequest, export_table_xlsx

        req = TableExportRequest(
            filename="도서별수불원장_91723",
            sheet_title="도서별수불원장",
            columns=[TableExportColumn(key="gdate", label="거래일자"), TableExportColumn(key="balance", label="현재고")],
            rows=[{"gdate": "2026.01.20", "balance": 54, "ignored": 1}, {"gdate": "합계", "balance": 914}],
        )
        res = await export_table_xlsx(req, current={"user_id": "t"})
        self.assertEqual(res.status_code, 200)
        self.assertIn("filename*=UTF-8''", res.headers["content-disposition"])
        ws = load_workbook(BytesIO(res.body)).active
        self.assertEqual([c.value for c in ws[1]], ["거래일자", "현재고"])
        self.assertEqual([c.value for c in ws[2]], ["2026.01.20", 54])
        self.assertEqual([c.value for c in ws[3]], ["합계", 914])

    async def test_row_cap(self) -> None:
        from fastapi import HTTPException

        from app.routers import export_table as m

        req = m.TableExportRequest(columns=[m.TableExportColumn(key="a", label="A")], rows=[{"a": 1}] * 3)
        m.EXPORT_TABLE_MAX_ROWS, saved = 2, m.EXPORT_TABLE_MAX_ROWS
        try:
            with self.assertRaises(HTTPException) as ctx:
                await m.export_table_xlsx(req, current={})
            self.assertEqual(ctx.exception.status_code, 413)
        finally:
            m.EXPORT_TABLE_MAX_ROWS = saved


if __name__ == "__main__":
    main()
