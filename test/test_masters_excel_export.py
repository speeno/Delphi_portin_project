"""기초관리 목록 일괄 엑셀(.xlsx) export 회귀.

대상: ``/api/v1/masters/exports/{customer,inbound-vendors,authors,book}.xlsx``
- masters_excel 워크북 빌더/페이지 수집 단위 검증(순수).
- 4개 엔드포인트: 200 + spreadsheet media-type + Content-Disposition + 본문 파싱.
- collect_all_rows 가 ceil=500 페이지를 끝까지 모으는지(>1 페이지) 검증.
서비스는 monkeypatch 로 DB 부작용 없이 가짜 페이지 데이터를 돌려준다.
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from unittest import TestCase, main
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "도서물류관리프로그램" / "backend"
sys.path.insert(0, str(BACKEND))

import asyncio  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.deps import get_user_context  # noqa: E402
from app.main import app  # noqa: E402
from app.routers.auth import get_current_user  # noqa: E402
from app.services import masters_excel, masters_service  # noqa: E402

_SID = "remote_138"
_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _user() -> dict:
    return {
        "user_id": "t_exp",
        "server_id": _SID,
        "role": "operator",
        "hcode": "BR01",
        "permissions": ["master.customer.read", "master.write"],
        "fxx_caps": {f"F1{n}": {"read": True, "write": True, "print": True} for n in (1, 2, 3, 4)},
    }


async def _override_user() -> dict:
    return _user()


async def _override_ctx() -> dict:
    u = _user()
    return {
        "user_id": u["user_id"],
        "server_id": u["server_id"],
        "role": u["role"],
        "hcode": u["hcode"],
        "branch_id": u["hcode"],
        "permissions": list(u["permissions"]),
        "tenant_id": "",
        "account_family": "",
        "active_build_id": "",
        "build_role": "",
        "account_type": "",
        "dist_hcode": "",
        "fxx_caps": dict(u["fxx_caps"]),
    }


def _make_fake_list(dataset: list[dict]):
    """list_* 서비스 대체 — keyword-only (server_id, limit, offset, ...) 시그니처 호환."""

    async def fake(*, server_id: str, limit: int, offset: int, **_kwargs):  # noqa: ARG001
        chunk = dataset[offset : offset + limit]
        has_more = (offset + len(chunk)) < len(dataset)
        return {
            "items": chunk,
            "page": {"limit": limit, "offset": offset, "total": len(dataset), "has_more": has_more},
        }

    return fake


def _positional(svc):
    """라우터의 fetch(limit, offset) 클로저와 동일하게 위치인자 → 서비스 호출로 어댑트."""

    async def fetch(limit: int, offset: int):
        return await svc(server_id="x", limit=limit, offset=offset)

    return fetch


def _read_sheet(content: bytes):
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    header = [c.value for c in ws[1]]
    return ws, header


class MastersExcelUnitTests(TestCase):
    def test_build_workbook_header_and_phone_merge(self) -> None:
        rows = [
            {"gcode": "A1", "gname": "가", "gtel1": "02", "gtel2": "1-2", "gpost": "01", "gjuso": "서울"},
            {"gcode": "A2", "gname": "나", "gtel1": "", "gtel2": "", "gpost": "", "gjuso": ""},
        ]
        data = masters_excel.build_list_workbook(
            sheet_title="거래처", columns=masters_excel.CUSTOMER_COLUMNS, rows=rows
        )
        ws, header = _read_sheet(data)
        self.assertEqual(header, ["코드", "거래처명", "전화", "우편번호", "주소"])
        self.assertEqual([c.value for c in ws[2]], ["A1", "가", "02-1-2", "01", "서울"])
        self.assertEqual(ws.max_row, 3)  # 헤더 + 2행

    def test_collect_all_rows_paginates_past_ceiling(self) -> None:
        dataset = [{"gcode": f"C{i:05d}", "gname": f"n{i}"} for i in range(1203)]
        rows, truncated = asyncio.run(
            masters_excel.collect_all_rows(_positional(_make_fake_list(dataset)), page_size=500)
        )
        self.assertFalse(truncated)
        self.assertEqual(len(rows), 1203)
        self.assertEqual(rows[0]["gcode"], "C00000")
        self.assertEqual(rows[-1]["gcode"], "C01202")

    def test_collect_all_rows_caps_and_flags_truncated(self) -> None:
        dataset = [{"gcode": str(i)} for i in range(50)]
        rows, truncated = asyncio.run(
            masters_excel.collect_all_rows(
                _positional(_make_fake_list(dataset)), page_size=500, max_rows=10
            )
        )
        self.assertTrue(truncated)
        self.assertEqual(len(rows), 10)


class MastersExcelExportRouterTests(TestCase):
    def setUp(self) -> None:
        self._prev_user = app.dependency_overrides.get(get_current_user)
        self._prev_ctx = app.dependency_overrides.get(get_user_context)
        app.dependency_overrides[get_current_user] = _override_user
        app.dependency_overrides[get_user_context] = _override_ctx
        self.client = TestClient(app)

    def tearDown(self) -> None:
        for dep, prev in ((get_current_user, self._prev_user), (get_user_context, self._prev_ctx)):
            if prev is not None:
                app.dependency_overrides[dep] = prev
            else:
                app.dependency_overrides.pop(dep, None)

    def _assert_xlsx(self, r) -> bytes:
        self.assertEqual(r.status_code, 200, r.text)
        self.assertEqual(r.headers["content-type"], _XLSX_MEDIA)
        self.assertIn("attachment", r.headers.get("content-disposition", ""))
        return r.content

    def _full_row(self, i: int) -> dict:
        return {
            "gcode": f"D{i:04d}", "ocode": "", "gname": f"거래처{i}", "gbun_name": "서점",
            "jubun": "서울", "gposa": "대표", "gnumb": "123", "guper": "도소매", "gjomo": "서적",
            "gpper": 0.0, "gpost": "01000", "gtel1": "02", "gtel2": f"{i:04d}",
            "gfax1": "", "gfax2": "", "gadd1": "주소", "gadd2": "",
            "grat1": 70.0, "grat2": 0, "grat3": 0, "grat4": 0, "grat5": 0, "grat6": 0,
            "gqut1": 0, "grat9": 0, "gbigo": "", "name1": "", "yesno": "", "pubun": "",
            "name2": "", "email": "", "gnum1": "",
        }

    def test_customer_export_full_columns_all_pages(self) -> None:
        dataset = [self._full_row(i) for i in range(1100)]  # 3 페이지(500/500/100)
        with patch.object(masters_service, "list_customer_master_full", side_effect=_make_fake_list(dataset)):
            r = self.client.get(f"/api/v1/masters/exports/customer.xlsx?serverId={_SID}")
        ws, header = _read_sheet(self._assert_xlsx(r))
        self.assertEqual(header[:3], ["거래처코드", "거래처코드2", "거래처명"])
        # DEC-149 — 담당자1(gpper 정본)·핸드폰번호(gphon)·한도액(gssum) 32→34.
        # DEC-234 — 전화/팩스 1·2 → 전화번호/팩스번호 합본(−2), 사업자주소 합본(+1) = 33. 헤더는 화면 라벨과 동일.
        self.assertEqual(len(header), 33)
        for h in ("전화번호", "팩스번호", "사업자등록번호", "사업자주소", "위탁", "기타"):
            self.assertIn(h, header)
        for h in ("전화번호1", "전화번호2", "사업자번호", "위탁공급율"):
            self.assertNotIn(h, header)
        self.assertEqual(ws.cell(row=2, column=header.index("전화번호") + 1).value, "02-0000")
        self.assertIn("담당자1", header)
        self.assertIn("핸드폰번호", header)
        self.assertIn("한도액", header)
        self.assertIn("대표자", header)  # Gposa=대표자 (의미 정정 반영)
        self.assertIn("업태", header)  # Guper=업태
        self.assertEqual(ws.max_row, 1101)  # 헤더 + 1100

    def test_customer_export_field_subset_in_catalog_order(self) -> None:
        dataset = [self._full_row(1)]
        with patch.object(masters_service, "list_customer_master_full", side_effect=_make_fake_list(dataset)):
            r = self.client.get(
                f"/api/v1/masters/exports/customer.xlsx?serverId={_SID}&fields=guper,gcode,gposa"
            )
        ws, header = _read_sheet(self._assert_xlsx(r))
        # 요청 순서와 무관하게 카탈로그 순서(거래처코드 → 대표자 → 업태).
        self.assertEqual(header, ["거래처코드", "대표자", "업태"])
        self.assertEqual([c.value for c in ws[2]], ["D0001", "대표", "도소매"])

    def test_customer_fields_catalog_endpoint(self) -> None:
        r = self.client.get("/api/v1/masters/exports/customer-fields")
        self.assertEqual(r.status_code, 200, r.text)
        fields = r.json()["fields"]
        by_key = {f["key"]: f["label"] for f in fields}
        self.assertEqual(by_key.get("gposa"), "대표자")
        self.assertEqual(by_key.get("guper"), "업태")
        self.assertEqual(by_key.get("gjomo"), "종목")

    def test_inbound_vendor_export(self) -> None:
        # DEC-172 — 헤더 = 입고처 세부내역 전면 카탈로그(거래처 DEC-149 동형, 목록 화면 순서).
        dataset = [{"gcode": "I1", "gname": "입고가", "gbun_name": "구분", "jubun": "서울",
                    "gposa": "대표", "guper": "업태", "gtel1": "031", "gtel2": "1", "gpost": "1",
                    "gadd1": "경기", "gadd2": "", "gpper": "홍길동", "gphon": "010-1", "gssum": 500}]
        with patch.object(masters_service, "list_inbound_vendors", side_effect=_make_fake_list(dataset)):
            r = self.client.get(f"/api/v1/masters/exports/inbound-vendors.xlsx?serverId={_SID}")
        ws, header = _read_sheet(self._assert_xlsx(r))
        self.assertEqual(header[:5], ["입고처구분", "입고처지역", "입고처코드", "입고처코드2", "입고처명"])
        for h in ("담당자", "핸드폰번호", "한도액", "한도", "계산서 거래처명", "정지사유", "전화번호", "주소1"):
            self.assertIn(h, header)
        self.assertNotIn("코드", header)  # 구 8컬럼 헤더 폐기(PK 헤더 = 입고처코드)
        self.assertEqual(ws.cell(row=2, column=header.index("담당자") + 1).value, "홍길동")
        self.assertEqual(ws.cell(row=2, column=header.index("한도액") + 1).value, 500)

    def test_inbound_vendor_export_fields_subset_keeps_pk(self) -> None:
        # DEC-172 — fields 선택 시 카탈로그 순서 유지 + PK(입고처코드) 항상 포함.
        dataset = [{"gcode": "I1", "gname": "입고가", "gpper": "홍길동"}]
        with patch.object(masters_service, "list_inbound_vendors", side_effect=_make_fake_list(dataset)):
            r = self.client.get(
                f"/api/v1/masters/exports/inbound-vendors.xlsx?serverId={_SID}&fields=gpper,gname"
            )
        _ws, header = _read_sheet(self._assert_xlsx(r))
        self.assertEqual(header, ["입고처코드", "입고처명", "담당자"])

    def test_inbound_vendor_fields_catalog(self) -> None:
        r = self.client.get("/api/v1/masters/exports/inbound-vendor-fields")
        self.assertEqual(r.status_code, 200, r.text)
        by_key = {f["key"]: f["label"] for f in r.json()["fields"]}
        self.assertEqual(by_key.get("gpper"), "담당자")
        self.assertEqual(by_key.get("gssum"), "한도액")
        self.assertEqual(by_key.get("gphon"), "핸드폰번호")
        self.assertEqual(by_key.get("grat7"), "한도")
        self.assertEqual(by_key.get("name2"), "계산서 거래처명")
        self.assertEqual(by_key.get("email"), "정지사유")

    def test_author_export(self) -> None:
        dataset = [{"gcode": "Z1", "gposa": "홍길동", "gbun_name": "소설", "gname": "직장",
                    "gtel1": "02", "gtel2": "9", "date1": "20260101", "gjice": "대표"}]
        with patch.object(masters_service, "list_authors", side_effect=_make_fake_list(dataset)):
            r = self.client.get(f"/api/v1/masters/exports/authors.xlsx?serverId={_SID}")
        _ws, header = _read_sheet(self._assert_xlsx(r))
        self.assertEqual(header, ["코드", "저자명", "저자구분", "직장명", "전화", "등록일자", "직책"])

    def test_book_export(self) -> None:
        # DEC-148 — 헤더 = 목록 화면 세부내역 컬럼 1:1 (gpost 라벨 "서가위치" 정정).
        dataset = [{"gcode": "B1", "gname": "책제목", "gjeja": "저자", "gisbn": "978",
                    "date1": "20260101", "gdang": 15000, "gpost": "H13, 25"}]
        with patch.object(masters_service, "list_books", side_effect=_make_fake_list(dataset)):
            r = self.client.get(f"/api/v1/masters/exports/book.xlsx?serverId={_SID}")
        ws, header = _read_sheet(self._assert_xlsx(r))
        self.assertEqual(
            header[:7],
            ["도서코드", "제목", "저자", "ISBN", "발행일", "단가", "서가위치"],
        )
        for h in ("도서분류", "판형", "원가", "위탁", "한도", "재고", "출고정지"):
            self.assertIn(h, header)
        self.assertNotIn("출판사", header)
        self.assertEqual(ws.cell(row=2, column=6).value, 15000)  # 단가 숫자 보존
        self.assertEqual(ws.cell(row=2, column=7).value, "H13, 25")


class MastersExcelStatic(TestCase):
    def test_routes_present_in_router(self) -> None:
        src = (BACKEND / "app" / "routers" / "masters.py").read_text(encoding="utf-8")
        for token in (
            '@router.get("/exports/customer.xlsx")',
            '@router.get("/exports/customer-fields")',
            '@router.get("/exports/inbound-vendors.xlsx")',
            '@router.get("/exports/authors.xlsx")',
            '@router.get("/exports/book.xlsx")',
        ):
            self.assertIn(token, src)


class MastersImportRouterTests(TestCase):
    def setUp(self) -> None:
        self._prev_user = app.dependency_overrides.get(get_current_user)
        self._prev_ctx = app.dependency_overrides.get(get_user_context)
        app.dependency_overrides[get_current_user] = _override_user
        app.dependency_overrides[get_user_context] = _override_ctx
        self.client = TestClient(app)

    def tearDown(self) -> None:
        for dep, prev in ((get_current_user, self._prev_user), (get_user_context, self._prev_ctx)):
            if prev is not None:
                app.dependency_overrides[dep] = prev
            else:
                app.dependency_overrides.pop(dep, None)

    def test_customer_import_round_trip_updates(self) -> None:
        rows = [
            {"gcode": "D0001", "gname": "가", "gposa": "대표가", "guper": "도소매", "grat1": 70},
            {"gcode": "D0002", "gname": "나", "gposa": "대표나", "guper": "소매", "grat1": 0},
        ]
        data = masters_excel.build_list_workbook(
            sheet_title="거래처", columns=masters_excel.CUSTOMER_FULL_COLUMNS, rows=rows
        )
        captured: list[tuple[str, dict]] = []

        async def fake_update(*, server_id, gcode, payload, scope_hcode=None):  # noqa: ANN001, ARG001
            captured.append((gcode, payload))
            return {"gcode": gcode, "updated_fields": list(payload.keys()), "updated_at": "t"}

        with patch.object(masters_service, "update_customer_master", side_effect=fake_update):
            r = self.client.post(
                f"/api/v1/masters/imports/customer.xlsx?serverId={_SID}",
                files={"file": ("customer_master.xlsx", data, _XLSX_MEDIA)},
            )
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertEqual(body["summary"]["total"], 2)
        self.assertEqual(body["summary"]["updated"], 2)
        self.assertEqual({g for g, _ in captured}, {"D0001", "D0002"})
        # PK(gcode)는 payload(수정필드)에 들어가지 않음 — 행 식별 전용.
        self.assertTrue(all("gcode" not in p for _, p in captured))
        # 대표자(gposa)·업태(guper) 가 역반영 payload 에 포함.
        d1 = dict(captured)["D0001"]
        self.assertEqual(d1.get("gposa"), "대표가")
        self.assertEqual(d1.get("guper"), "도소매")

    def test_import_bad_header_returns_400(self) -> None:
        data = masters_excel.build_list_workbook(
            sheet_title="x", columns=[("엉뚱헤더", "zzz")], rows=[{"zzz": "1"}]
        )
        r = self.client.post(
            f"/api/v1/masters/imports/book.xlsx?serverId={_SID}",
            files={"file": ("x.xlsx", data, _XLSX_MEDIA)},
        )
        self.assertEqual(r.status_code, 400, r.text)

    def test_import_rejects_non_xlsx(self) -> None:
        r = self.client.post(
            f"/api/v1/masters/imports/authors.xlsx?serverId={_SID}",
            files={"file": ("x.csv", b"a,b,c", "text/csv")},
        )
        self.assertEqual(r.status_code, 400, r.text)


class OrderByClauseTests(TestCase):
    def test_valid_key_with_secondary(self) -> None:
        self.assertEqual(
            masters_service._order_by_clause(
                "gname", "desc", allowed=masters_service._CUSTOMER_SORTS, default="Gcode"
            ),
            "ORDER BY Gname DESC, Gcode",
        )

    def test_injection_key_falls_back_to_default(self) -> None:
        self.assertEqual(
            masters_service._order_by_clause(
                "Gcode; DROP TABLE x", "asc", allowed=masters_service._CUSTOMER_SORTS, default="Gcode"
            ),
            "ORDER BY Gcode",
        )

    def test_default_key_no_duplicate_secondary(self) -> None:
        self.assertEqual(
            masters_service._order_by_clause(
                "gcode", "asc", allowed=masters_service._CUSTOMER_SORTS, default="Gcode"
            ),
            "ORDER BY Gcode ASC",
        )


class BookDateFilterTests(TestCase):
    def test_date_filter_normalizes_separators_and_excludes_empty(self) -> None:
        captured: dict = {}

        async def fake_eq(server_id, sql, params=()):  # noqa: ANN001, ARG001
            if "row_count" in sql:
                return [{"row_count": 0}]
            captured["sql"] = sql
            captured["params"] = params
            return []

        with patch.object(masters_service, "execute_query", side_effect=fake_eq):
            asyncio.run(
                masters_service.list_books(
                    server_id=_SID,
                    date_from="2026-01-01",
                    date_to="2026-12-31",
                    scope_hcode="BR01",
                )
            )
        sql = captured["sql"]
        # 구분자 제거 정규화 비교 + 빈 Date1 제외.
        self.assertIn("REPLACE(REPLACE(REPLACE(IFNULL(Date1", sql)
        self.assertIn("<> ''", sql)
        self.assertIn("20260101", captured["params"])
        self.assertIn("20261231", captured["params"])


if __name__ == "__main__":
    main()
