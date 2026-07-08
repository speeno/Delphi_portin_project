"""DEC-092 — 전자책 판매분석 (Web_Ebook_Sales) 회귀 가드.

축
--
1) 서비스: upsert/삭제/목록 SQL 이 hcode 스코프 + mysql3-safe(REPLACE INTO) 인지.
2) 업로드 파서: 롱 포맷(년월|판매처|팀|부수|금액) 헤더 검증 + 동일 키 합산 + 별칭/불량행.
3) 보고서 빌더: 서식 2종(연간/월범위) 피벗 — 채널 순서, 1팀/2팀/계, 전년대비, 계 행.
4) 라우터: /api/v1/stats/ebook-sales 6종 OpenAPI 등록.
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path
from unittest import IsolatedAsyncioTestCase, TestCase, main
from unittest.mock import patch

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "도서물류관리프로그램" / "backend"
sys.path.insert(0, str(BACKEND))

from app.services import ebook_sales_service as svc  # noqa: E402


def _xlsx(rows: list[list], headers: list[str] | None = None) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers or ["년월", "판매처", "팀", "부수", "금액"])
    for r in rows:
        ws.append(r)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class ServiceSqlTests(IsolatedAsyncioTestCase):
    def setUp(self) -> None:
        svc.clear_ensured_for_tests()

    async def _run(self, coro_fn, captured, **kw):
        async def fake_execute(server_id, sql, params=()):  # noqa: ARG001
            captured.append((sql, params))
            up = sql.strip().upper()
            if up.startswith("SELECT COUNT("):
                return [{"cnt": 1}]
            if "SUM(QTY)" in up:
                return [{"q": 10, "a": 150000}]
            if up.startswith("SELECT DISTINCT"):
                return [{"Channel": "교보"}, {"Channel": "스콘"}, {"Channel": "아카디피아"}]
            if up.startswith("SELECT YM"):
                return [{"Ym": "202601", "Channel": "교보", "Team": "1", "Qty": 10, "Amt": 150000}]
            return []

        with patch.object(svc, "execute_query", side_effect=fake_execute):
            return await coro_fn(**kw)

    async def test_upsert_uses_replace_into_with_hcode(self) -> None:
        cap: list = []
        res = await self._run(
            svc.upsert_row, cap, server_id="remote_138", scope_hcode="5019",
            ym="2026-01", channel="교보", team="1팀", qty=10, amt=150000,
        )
        self.assertEqual(res["action"], "upserted")
        self.assertEqual(res["ym"], "202601")
        self.assertEqual(res["team"], "1")
        replace = next(s for s, _ in cap if s.strip().upper().startswith("REPLACE INTO"))
        self.assertIn("Web_Ebook_Sales", replace)
        _, params = next((s, p) for s, p in cap if s.strip().upper().startswith("REPLACE INTO"))
        self.assertEqual(params[0], "5019")  # hcode 스코프 태깅

    async def test_upsert_zero_zero_deletes(self) -> None:
        cap: list = []
        res = await self._run(
            svc.upsert_row, cap, server_id="remote_138", scope_hcode="5019",
            ym="202601", channel="교보", team="2", qty=0, amt=0,
        )
        self.assertEqual(res["action"], "deleted")
        self.assertTrue(any(s.strip().upper().startswith("DELETE FROM") for s, _ in cap))

    async def test_upsert_validates_ym_and_team(self) -> None:
        with self.assertRaises(svc.EbookSalesValidationError):
            await self._run(
                svc.upsert_row, [], server_id="remote_138", scope_hcode="5019",
                ym="2026-13", channel="교보", team="1", qty=1, amt=1,
            )
        with self.assertRaises(svc.EbookSalesValidationError):
            await self._run(
                svc.upsert_row, [], server_id="remote_138", scope_hcode="5019",
                ym="202601", channel="교보", team="3팀", qty=1, amt=1,
            )

    async def test_list_scopes_by_hcode_and_orders_channels(self) -> None:
        cap: list = []
        res = await self._run(
            svc.list_rows, cap, server_id="remote_138", scope_hcode="5019",
            ym_from="2026-01", ym_to="2026-03",
        )
        list_sql, list_params = next(
            (s, p) for s, p in cap if s.strip().upper().startswith("SELECT YM")
        )
        self.assertIn("Hcode = %s", list_sql)
        self.assertEqual(list_params[0], "5019")
        self.assertEqual(res["items"][0]["ym"], "202601")
        # 채널 순서: 서식 우선순위(아카디피아, 교보) → 나머지(스콘)
        self.assertEqual(res["channels"], ["아카디피아", "교보", "스콘"])
        self.assertEqual(res["totals"], {"qty": 10, "amt": 150000})


class UploadParserTests(TestCase):
    def test_parse_ok_and_merges_same_key(self) -> None:
        data = _xlsx([
            ["2026-01", "교보", "1팀", 3, 30000],
            ["2026.01", "교보", "1", 2, 20000],   # 동일 키 → 합산
            ["202602", "스콘", "2팀", 5, 75000],
        ])
        parsed = svc.parse_upload_xlsx(data)
        self.assertTrue(parsed["header_ok"])
        rows = {(r["ym"], r["channel"], r["team"]): r for r in parsed["rows"]}
        self.assertEqual(rows[("202601", "교보", "1")]["qty"], 5)
        self.assertEqual(rows[("202601", "교보", "1")]["amt"], 50000)
        self.assertEqual(rows[("202602", "스콘", "2")]["qty"], 5)

    def test_parse_bad_header_rejected(self) -> None:
        parsed = svc.parse_upload_xlsx(_xlsx([], headers=["연도", "채널", "팀", "부수", "금액"]))
        self.assertFalse(parsed["header_ok"])

    def test_parse_skips_bad_rows_with_warning(self) -> None:
        parsed = svc.parse_upload_xlsx(_xlsx([
            ["2026-01", "교보", "3팀", 1, 100],   # 팀 불량
            ["", "", "", "", ""],                  # 빈 행 무시
            ["2026-02", "노팅", "2", 1, 100],
        ]))
        self.assertTrue(parsed["header_ok"])
        self.assertEqual(len(parsed["rows"]), 1)
        self.assertTrue(parsed["warnings"])


class ReportBuilderTests(TestCase):
    _ROWS = [
        {"ym": "202501", "channel": "교보", "team": "1", "qty": 3, "amt": 30000},
        {"ym": "202502", "channel": "교보", "team": "2", "qty": 2, "amt": 20000},
        {"ym": "202507", "channel": "교보", "team": "1", "qty": 7, "amt": 70000},  # 월범위 밖
        {"ym": "202601", "channel": "교보", "team": "1", "qty": 5, "amt": 50000},
        {"ym": "202601", "channel": "스콘", "team": "1", "qty": 1, "amt": 9000},
    ]

    def test_two_sheets_with_layout(self) -> None:
        from openpyxl import load_workbook

        content = svc.build_report_xlsx(self._ROWS, month_from=1, month_to=3)
        wb = load_workbook(BytesIO(content))
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn("연간", wb.sheetnames[0])
        self.assertIn("1월~3월", wb.sheetnames[1])

    def test_month_range_pivot_values_and_yoy(self) -> None:
        from openpyxl import load_workbook

        content = svc.build_report_xlsx(self._ROWS, month_from=1, month_to=3)
        wb = load_workbook(BytesIO(content))
        ws = wb[wb.sheetnames[1]]  # 1월~3월 시트
        grid = {}
        for row in ws.iter_rows(values_only=False):
            for c in row:
                if c.value is not None:
                    grid[(c.row, c.column)] = c.value
        vals = list(grid.values())
        # 헤더 요소
        self.assertIn("연도", vals)
        self.assertIn("교보", vals)
        self.assertIn("전년대비", vals)
        self.assertIn("*** 부수 ***", vals)
        self.assertIn("*** 금액 ***", vals)
        # 부수 블록: 2026 교보 1팀=5 (1~3월), 2025 교보 1팀=3/2팀=2 계=5.
        # 헤더에서 교보 그룹 시작 컬럼과 연도 행을 찾아 검증.
        kyobo_col = next(col for (r, col), v in grid.items() if v == "교보")
        header_row = next(r for (r, col), v in grid.items() if v == "교보" and col == kyobo_col)
        y2026 = next(r for (r, c), v in grid.items() if v == 2026 and c == 1 and r > header_row)
        y2025 = next(r for (r, c), v in grid.items() if v == 2025 and c == 1 and r > header_row)
        self.assertEqual(grid.get((y2026, kyobo_col)), 5)          # 1팀
        self.assertEqual(grid.get((y2026, kyobo_col + 2)), 5)      # 계
        self.assertEqual(grid.get((y2026, kyobo_col + 3)), 0)      # 전년대비 = 5-5
        self.assertEqual(grid.get((y2025, kyobo_col)), 3)
        self.assertEqual(grid.get((y2025, kyobo_col + 1)), 2)
        self.assertEqual(grid.get((y2025, kyobo_col + 2)), 5)
        # 2025 는 전년(2024) 데이터 없음 → 전년대비 공란.
        self.assertIsNone(grid.get((y2025, kyobo_col + 3)))

    def test_annual_sheet_includes_out_of_range_months(self) -> None:
        from openpyxl import load_workbook

        content = svc.build_report_xlsx(self._ROWS, month_from=1, month_to=3)
        wb = load_workbook(BytesIO(content))
        ws = wb[wb.sheetnames[0]]  # 연간
        grid = {(c.row, c.column): c.value for row in ws.iter_rows() for c in row if c.value is not None}
        kyobo_col = next(col for (r, col), v in grid.items() if v == "교보")
        y2025 = next(r for (r, c), v in grid.items() if v == 2025 and c == 1)
        # 연간: 2025 교보 1팀 = 3(1월) + 7(7월) = 10.
        self.assertEqual(grid.get((y2025, kyobo_col)), 10)


class RouterRegistrationTests(TestCase):
    def test_openapi_routes_registered(self) -> None:
        from app.main import app

        paths = {getattr(r, "path", "") for r in app.routes}
        for p in (
            "/api/v1/stats/ebook-sales",
            "/api/v1/stats/ebook-sales/template.xlsx",
            "/api/v1/stats/ebook-sales/import.xlsx",
            "/api/v1/stats/ebook-sales/export.xlsx",
        ):
            self.assertIn(p, paths, f"missing route: {p}")


if __name__ == "__main__":
    main()
