#!/usr/bin/env python3
"""레거시 마스터 엑셀(거래처/도서) → parity baseline JSON 정규화.

목적
----
레거시 델파이에서 특정 테넌트(예: 교문사) 계정으로 조회·내보낸 거래처/도서
리스트 엑셀을 ``debug/baselines/<slug>_customers.json`` / ``<slug>_books.json``
으로 정규화한다. 이 baseline 은 [`probe_tenant_master_parity.py`](probe_tenant_master_parity.py)
가 웹 API 응답(`/api/v1/masters/customer`·`/book`)과 diff 하는 Ground Truth 다.

계정 무관(SRP/OCP)
------------------
- 교문사 전용 컬럼 가정 없이, 헤더 라벨을 정규 키로 매핑한다.
- 타 계정 엑셀도 동일 헤더 규약(거래처=코드/거래처명, 도서=도서코드/도서명)이면
  ``--slug <tenant>`` 만 바꿔 재사용한다.

비밀 정책
---------
- 엑셀에는 자격증명이 없으며, 본 스크립트도 자격증명을 읽거나 저장하지 않는다.

사용
----
    python3 debug/import_legacy_master_baseline.py \
        --slug gyomunsa \
        --customers "~/Downloads/교문사 거래처관리 리스트(위러브솔루션).xlsx" \
        --books "~/Downloads/교문사 도서관리 리스트(위러브솔루션).xlsx"
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
BASELINE_DIR = REPO_ROOT / "debug" / "baselines"

# 레거시 헤더 라벨 → 정규 키. 존재하는 컬럼만 채택 (없으면 skip).
_CUSTOMER_COLS = {
    "코드": "gcode",
    "거래처명": "gname",
    "구분": "gubun",
    "지역": "region",
    "대표자명": "owner",
    "사업자번호": "biz_no",
    "전화번호": "tel",
    "우편번호": "post",
    "주소": "addr",
}
_BOOK_COLS = {
    "도서코드": "gcode",
    "도서명": "gname",
    "도서분류": "category",
    "도서처리": "status",
    "저자명": "author",
    "단가": "price",
    "ISBN번호": "isbn",
}


def _norm_cell(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return v


def _norm_gcode(v: Any) -> str | None:
    """Gcode 정규화 — 엑셀이 int/float/str 혼용이므로 문자열로 통일."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return s or None


def _load_sheet(path: Path, col_map: dict[str, str]) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    # 헤더 라벨 → 컬럼 인덱스 (공백 정규화)
    idx_of: dict[str, int] = {}
    for i, h in enumerate(header):
        if isinstance(h, str):
            idx_of[h.strip()] = i
    take = {key: idx_of[label] for label, key in col_map.items() if label in idx_of}

    items: list[dict[str, Any]] = []
    by_gcode: dict[str, str] = {}
    for r in rows:
        if not any(c is not None for c in r):
            continue
        rec: dict[str, Any] = {}
        for key, ci in take.items():
            raw = r[ci] if ci < len(r) else None
            rec[key] = _norm_gcode(raw) if key == "gcode" else _norm_cell(raw)
        gc = rec.get("gcode")
        if gc is None:
            continue
        items.append(rec)
        if "gname" in rec and isinstance(rec["gname"], str):
            by_gcode.setdefault(gc, rec["gname"])
    wb.close()

    return {
        "count": len(items),
        "columns": sorted(take.keys()),
        "by_gcode_name": by_gcode,
        "items": items,
    }


def _write(slug: str, kind: str, payload: dict[str, Any], source: Path) -> Path:
    BASELINE_DIR.mkdir(parents=True, exist_ok=True)
    out = BASELINE_DIR / f"{slug}_{kind}.json"
    doc = {
        "schema_version": "1.0.0",
        "tenant_slug": slug,
        "kind": kind,
        "source_file": source.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "note": "레거시 조회 결과 Ground Truth — 자격증명 0건.",
        **payload,
    }
    tmp = out.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(out)
    return out


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n", 1)[0])
    p.add_argument("--slug", required=True, help="테넌트 슬러그 (예: gyomunsa)")
    p.add_argument("--customers", help="거래처 엑셀 경로")
    p.add_argument("--books", help="도서 엑셀 경로")
    args = p.parse_args(argv)

    if not args.customers and not args.books:
        p.error("--customers 또는 --books 중 최소 하나는 필요")

    if args.customers:
        cp = Path(os.path.expanduser(args.customers))
        if not cp.exists():
            print(f"[ERR] customers not found: {cp}")
            return 2
        payload = _load_sheet(cp, _CUSTOMER_COLS)
        out = _write(args.slug, "customers", payload, cp)
        print(f"[OK] customers count={payload['count']} → {out}")

    if args.books:
        bp = Path(os.path.expanduser(args.books))
        if not bp.exists():
            print(f"[ERR] books not found: {bp}")
            return 2
        payload = _load_sheet(bp, _BOOK_COLS)
        out = _write(args.slug, "books", payload, bp)
        print(f"[OK] books count={payload['count']} → {out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
